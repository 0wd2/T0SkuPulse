import streamlit as st
import plotly.graph_objects as go
import polars as pl
import pandas as pd
from io import BytesIO
from datetime import date, timedelta
from openpyxl import load_workbook
from plotly.subplots import make_subplots
import plotly.express as px
import numpy as np
import warnings
import fastexcel
warnings.filterwarnings("ignore")
import io
import time
from xlsxwriter import Workbook

# 1、设置页面标题
st.set_page_config(page_title="T0SKU指标监控", layout="wide", page_icon="📊")

# st.markdown("""
#     <style>
#     /* 强制固定顶部容器 */
#     [data-testid="stVerticalBlock"] > div:has(#fixed-header-anchor) {
#         position: sticky;
#         top: 2.8rem;
#         background-color: white;
#         z-index: 1000;
#         border-bottom: 2px solid #f0f2f6;
#         padding-bottom: 10px;
#     }
#     #fixed-header-anchor { display: none; }

#     /* 指标卡美化 */
#     div[data-testid="metric-container"] {
#         background-color: #f8f9fa;
#         border: 1px solid #eef0f2;
#         padding: 10px;
#         border-radius: 8px;
#     }
#     </style>
#     """, unsafe_allow_html=True)

# 2、初始化 Session State
if "df_fahuo" not in st.session_state: st.session_state.df_fahuo = None
if "df_yuce" not in st.session_state: st.session_state.df_yuce = None
if "df_ganyu" not in st.session_state: st.session_state.df_ganyu = None
if "filter_ver" not in st.session_state: st.session_state.filter_ver = 0
if "t0_date" not in st.session_state: st.session_state.t0_date = None
if "fahuo_frequency" not in st.session_state: st.session_state.fahuo_frequency = None
if "target_week" not in st.session_state: st.session_state.target_week = None
if "committed_filters" not in st.session_state:
    st.session_state.committed_filters = {"sub_market": [], "category": [], "mrpsku": []}
if "stock_turnover" not in st.session_state: st.session_state.stock_turnover = None
if "filter_market" not in st.session_state: st.session_state.filter_market = None
if "df_country_turnover" not in st.session_state: st.session_state.df_country_turnover = None
if "df_stock_turnover" not in st.session_state: st.session_state.df_stock_turnover = None
if "df_历史海外周转" not in st.session_state: st.session_state.df_历史海外周转 = None
if "df_LT前预测偏差" not in st.session_state: st.session_state.df_LT前预测偏差 = None
if "df_断货无在途" not in st.session_state: st.session_state.df_断货无在途 = None

st.markdown("""
        <style>
        /* 1. 强制放大 Label (指标名称) */
        [data-testid="stMetricLabel"] div {
            font-size: 24px !important;
            font-weight: bold !important;
            font-family: "Microsoft YaHei", sans-serif;
            color: #31333F !important; /* 确保颜色清晰 */
        }

        /* 2. 强制放大 Delta (下方目标/变动文字) */
        [data-testid="stMetricDelta"] div {
            font-size: 19px !important;
        }
        </style>
        """, unsafe_allow_html=True)

@st.cache_data(show_spinner="正在解析并格式化数据，请稍候...", ttl=3600)
def process_uploaded_files(uploaded_files):
    """
    处理所有上传的文件并返回一个字典，存储各业务数据框。
    使用缓存避免重复读取大文件。
    """
    data_pool = {
        "df_fahuo": None, "df_yuce": None, "df_ganyu": None,
        "df_stock_turnover": None, "df_country_turnover": None,
        "df_历史海外周转": None, "df_LT前预测偏差": None,
        "df_断货无在途": None, "stock_turnover": None
    }

    for uploaded_file in uploaded_files:
        file_name = uploaded_file.name
        file_bytes = uploaded_file.getvalue()
        file_type = file_name.split('.')[-1].lower()

        try:
            if file_type == 'xlsx':
                # 使用 BytesIO 读取
                with io.BytesIO(file_bytes) as f:
                    # 获取所有 Sheet 名称
                    workbook = load_workbook(f, read_only=True)
                    sn = workbook.sheetnames
                    
                    # 定义内部格式化工具
                    def format_df(df):
                        if df is not None and not df.empty:
                            if '日期' in df.columns:
                                df['日期'] = pd.to_datetime(df['日期'])
                            return df
                        return None

                    # 批量读取各 Sheet
                    if '发货指标' in sn: 
                        data_pool["df_fahuo"] = format_df(pl.read_excel(f, sheet_name='发货指标').to_pandas())
                    if '预测偏差' in sn: 
                        data_pool["df_yuce"] = format_df(pl.read_excel(f, sheet_name='预测偏差').to_pandas())
                    if '干预偏差' in sn: 
                        data_pool["df_ganyu"] = format_df(pl.read_excel(f, sheet_name='干预偏差').to_pandas())
                    if '库存与周转' in sn: 
                        data_pool["df_stock_turnover"] = format_df(pl.read_excel(f, sheet_name='库存与周转').to_pandas())
                    if '国内库存周转' in sn: 
                        data_pool["df_country_turnover"] = format_df(pl.read_excel(f, sheet_name='国内库存周转').to_pandas())
                    if '历史海外周转' in sn: 
                        data_pool["df_历史海外周转"] = format_df(pl.read_excel(f, sheet_name='历史海外周转').to_pandas())
                    if 'LT前预测偏差' in sn: 
                        data_pool["df_LT前预测偏差"] = format_df(pl.read_excel(f, sheet_name='LT前预测偏差').to_pandas())
                    if '断货无在途' in sn: 
                        data_pool["df_断货无在途"] = format_df(pl.read_excel(f, sheet_name='断货无在途').to_pandas())
            
            elif file_type == 'parquet':
                with io.BytesIO(file_bytes) as f:
                    data_pool["stock_turnover"] = pd.read_parquet(f)
                    
        except Exception as e:
            st.error(f"解析文件 {file_name} 时出错: {e}")
            
    return data_pool




# 3、侧边栏配置
with st.sidebar:
    st.header("⚙️ 配置参数")
    today = date.today()
    default_monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
    t0_date_val = st.date_input("🗓️ 当前周周一", value=default_monday)
    st.session_state.t0_date = t0_date_val

    up_folder_btn = st.file_uploader("请选择上传文件：", type=['xlsx','parquet'], accept_multiple_files=True)
    
    if up_folder_btn:
        # 调用缓存函数处理文件
        with st.spinner("正在快速加载数据..."):
            processed_data = process_uploaded_files(up_folder_btn)
            
            # 将处理后的数据更新到 session_state（仅在有值时更新，避免覆盖已有的其他数据）
            for key, df in processed_data.items():
                if df is not None:
                    st.session_state[key] = df
            
            st.success("✅ 数据加载完成 (已缓存)")
    
    st.divider()

    # --- 第三部分：目录大纲 ---
    st.sidebar.subheader("📌 快速导航")

    # 使用 HTML 和 CSS 打造更具设计感的菜单
    toc_html = """
    <style>
        .toc-container {
            padding: 10px 0;
        }
        .toc-item {
            display: flex;
            align-items: center;
            padding: 12px 15px;
            margin: 8px 0;
            background-color: #f8f9fa;
            border-radius: 10px;
            text-decoration: none;
            color: #31333F !important;
            font-weight: 500;
            transition: all 0.3s ease;
            border: 1px solid #e0e0e0;
        }
        .toc-item:hover {
            background-color: #ffffff;
            border-color: #ff4b4b;
            box-shadow: 0 4px 6px rgba(0,0,0,0.05);
            transform: translateX(5px);
            color: #ff4b4b !important;
        }
        .toc-icon {
            margin-right: 12px;
            font-size: 1.2rem;
        }
    </style>
    <div class="toc-container">
        <a href="#1" class="toc-item">
            <span class="toc-icon">📉</span> 海外库存周转
        </a>
        <a href="#2" class="toc-item">
            <span class="toc-icon">🚚</span> 发货指标监控
        </a>
        <a href="#3" class="toc-item">
            <span class="toc-icon">📈</span> 预测偏差
        </a>
        <a href="#4" class="toc-item">
            <span class="toc-icon">🛠️</span> 干预偏差
        </a>
    </div>
    """
    st.sidebar.markdown(toc_html, unsafe_allow_html=True)

# =========================================================
# 4、固定区域：Head、KPI卡片、筛选面板
# =========================================================
fixed_container = st.container()

with fixed_container:
    # 锚点
    # st.markdown('<div id="fixed-header-anchor"></div>', unsafe_allow_html=True)
    st.markdown("<h2 style='text-align: center; color: #1E3A8A; margin:0;'>📊 T0SKU 全链路指标监控</h2>",
                unsafe_allow_html=True)
    # 显示当前周周一,把日期变为年周
    current_week = st.session_state.t0_date.strftime("%Yw%V")
    # 年份只显示后两位，yyWww
    current_week = current_week[2:]
    head_1,head_2 = st.columns([2,3])
    with head_1:
        st.markdown(f"## 指标观测时间：{current_week}({st.session_state.t0_date.strftime('%m-%d')})")
    with head_2:
        toc_html = """
            <style>
                .toc-container {
                    display: flex;           /* 开启弹性布局 */
                    flex-direction: row;     /* 横向排列 */
                    flex-wrap: wrap;         /* 换行排版，防止手机端溢出 */
                    gap: 12px;               /* 按钮之间的间距 */
                    padding: 10px 0;
                    justify-content: flex-start; /* 靠左对齐，也可改为 space-between */
                }
                .toc-item {
                    display: flex;
                    align-items: center;
                    justify-content: center; /* 内容居中 */
                    padding: 8px 16px;       /* 调整内边距，使其更像水平标签 */
                    background-color: #f8f9fa;
                    border-radius: 8px;
                    text-decoration: none;
                    color: #31333F !important;
                    font-weight: 500;
                    font-size: 14px;         /* 稍微缩小字体以适应横排 */
                    transition: all 0.2s ease;
                    border: 1px solid #e0e0e0;
                    white-space: nowrap;     /* 禁止文字换行 */
                    flex: 1;                 /* 选填：让所有按钮等宽 */
                    min-width: 120px;        /* 按钮最小宽度 */
                    max-width: 200px;        /* 按钮最大宽度 */
                }
                .toc-item:hover {
                    background-color: #ffffff;
                    border-color: #ff4b4b;
                    box-shadow: 0 4px 8px rgba(255, 75, 75, 0.1);
                    transform: translateY(-2px); /* 悬停改为向上微动，横向更自然 */
                    color: #ff4b4b !important;
                }
                .toc-icon {
                    margin-right: 8px;
                    font-size: 1.1rem;
                }
                /* 适配移动端：如果屏幕太窄，自动变成两行 */
                @media (max-width: 768px) {
                    .toc-item {
                        flex: 1 1 45%; 
                    }
                }
            </style>
            <div class="toc-container">
                <a href="#0" class="toc-item">
                    <span class="toc-icon">📊</span> 指标概况
                </a>
                <a href="#1" class="toc-item">
                    <span class="toc-icon">📉</span> 海外库存周转
                </a>
                <a href="#2" class="toc-item">
                    <span class="toc-icon">🚚</span> 发货监控
                </a>
                <a href="#3" class="toc-item">
                    <span class="toc-icon">📈</span> 预测偏差
                </a>
                <a href="#4" class="toc-item">
                    <span class="toc-icon">🛠️</span> 干预偏差
                </a>
            </div>
            """
        st.markdown(toc_html, unsafe_allow_html=True)

    # 筛选面板逻辑
    if st.session_state.df_fahuo is not None:
        df_ref = st.session_state.df_fahuo
        ver = st.session_state.filter_ver
        c1, c2, c3, c4, c5 = st.columns([3, 3, 3, 1, 1], vertical_alignment="bottom")
        with c1:
            ui_market = st.multiselect("🌍 子市场", sorted(df_ref['子市场'].unique().tolist()), key=f"m_{ver}")
            st.session_state.committed_filters["sub_market"] = ui_market
        with c2:
            cat_opts = df_ref[df_ref['子市场'].isin(ui_market)]['品类'].unique() if ui_market else df_ref[
                '品类'].unique()
            ui_category = st.multiselect("📦 品类", sorted(cat_opts.tolist()), key=f"c_{ver}")
            st.session_state.committed_filters["category"] = ui_category
        with c3:
            sku_opts = df_ref[df_ref['品类'].isin(ui_category)]['主料mrpsku'].unique() if ui_category else df_ref[
                '主料mrpsku'].unique()
            ui_sku = st.multiselect("🔑 MRPSKU", sorted(sku_opts.tolist()), key=f"s_{ver}")
            st.session_state.committed_filters["mrpsku"] = ui_sku
        # with c4:
        #     if st.button("🚀 确认", width='stretch', type="primary"):
        #         st.session_state.committed_filters = {"sub_market": ui_market, "category": ui_category,
        #                                               "mrpsku": ui_sku}
        #         st.rerun()
        # with c5:
        #     if st.button("重置", width='stretch'):
        #         st.session_state.filter_ver += 1
        #         st.session_state.committed_filters = {"sub_market": [], "category": [], "mrpsku": []}
        #         st.rerun()
                
    if st.session_state.df_stock_turnover is not None:
        filters = st.session_state.committed_filters
        df_stock_turnover_kpi = st.session_state.df_stock_turnover.copy()
        if filters["sub_market"]: df_stock_turnover_kpi = df_stock_turnover_kpi[df_stock_turnover_kpi["子市场"].isin(filters["sub_market"])]
        if filters["category"]: df_stock_turnover_kpi = df_stock_turnover_kpi[df_stock_turnover_kpi["品类"].isin(filters["category"])]
        if filters["mrpsku"]: df_stock_turnover_kpi = df_stock_turnover_kpi[df_stock_turnover_kpi["主料mrpsku"].isin(filters["mrpsku"])]

        df_yuce_kpi = st.session_state.df_yuce.copy()
        if filters["sub_market"]: df_yuce_kpi = df_yuce_kpi[df_yuce_kpi["子市场"].isin(filters["sub_market"])]
        if filters["category"]: df_yuce_kpi = df_yuce_kpi[df_yuce_kpi["品类"].isin(filters["category"])]
        if filters["mrpsku"]: df_yuce_kpi = df_yuce_kpi[df_yuce_kpi["主料mrpsku"].isin(filters["mrpsku"])]

        df_LT前预测偏差 = st.session_state.df_LT前预测偏差.copy()
        if filters["sub_market"]: df_LT前预测偏差 = df_LT前预测偏差[df_LT前预测偏差["子市场"].isin(filters["sub_market"])]
        if filters["category"]: df_LT前预测偏差 = df_LT前预测偏差[df_LT前预测偏差["品类"].isin(filters["category"])]
        if filters["mrpsku"]: df_LT前预测偏差 = df_LT前预测偏差[df_LT前预测偏差["主料mrpsku"].isin(filters["mrpsku"])]

        df_ganyu_kpi = st.session_state.df_ganyu.copy()
        if filters["sub_market"]: df_ganyu_kpi = df_ganyu_kpi[df_ganyu_kpi["子市场"].isin(filters["sub_market"])]
        if filters["category"]: df_ganyu_kpi = df_ganyu_kpi[df_ganyu_kpi["品类"].isin(filters["category"])]
        if filters["mrpsku"]: df_ganyu_kpi = df_ganyu_kpi[df_ganyu_kpi["主料mrpsku"].isin(filters["mrpsku"])]

        # df_ganyu_bi = st.session_state.df_ganyu_bi.copy()
        
        df_country_turnover = st.session_state.df_country_turnover.copy()
        # if filters["sub_market"]: df_country_turnover = df_country_turnover[df_country_turnover["子市场"].isin(filters["sub_market"])]
        # if filters["category"]: df_country_turnover = df_country_turnover[df_country_turnover["品类"].isin(filters["category"])]
        # if filters["mrpsku"]: df_country_turnover = df_country_turnover[df_country_turnover["主料mrpsku"].isin(filters["mrpsku"])]

        # 计算历史海外在库在途周转
        df_历史海外周转 = st.session_state.df_历史海外周转.copy()
        if filters["sub_market"]: df_历史海外周转 = df_历史海外周转[df_历史海外周转["子市场"].isin(filters["sub_market"])]
        if filters["category"]: df_历史海外周转 = df_历史海外周转[df_历史海外周转["品类"].isin(filters["category"])]
        if filters["mrpsku"]: df_历史海外周转 = df_历史海外周转[df_历史海外周转["主料mrpsku"].isin(filters["mrpsku"])]


        with st.container(border=True):
            st.markdown("**🎯 指标概况**")
            last_dt_history = st.session_state.t0_date.strftime("%Yw%V")
            curr_avg_yuce = df_yuce_kpi[(df_yuce_kpi['周数']==last_dt_history)]['单周预测偏差率'].abs().mean()
            curr_avg_huanbiyuce = df_yuce_kpi[(df_yuce_kpi['周数']==last_dt_history)]['环比预测偏差率'].abs().mean()
            if len(df_LT前预测偏差[(df_LT前预测偏差['周数']==last_dt_history)]) > 0:
                curr_avg_ltyuce = df_LT前预测偏差[(df_LT前预测偏差['周数']==last_dt_history)]['实际周销与预测周销的预测偏差率'].abs().mean()
            else:
                curr_avg_ltyuce = 0
            # ganyu_intervention_rate = df_ganyu_bi["有干预样本数"].sum() / df_ganyu_bi["总样本数"].sum()
            ganyu_intervention_rate = len(df_ganyu_kpi[(df_ganyu_kpi['周数']==last_dt_history) & (df_ganyu_kpi['是否有干预']=="是")]) / len(df_ganyu_kpi[df_ganyu_kpi['周数']==last_dt_history])
            curr_avg_ganyu = df_ganyu_kpi[(df_ganyu_kpi['周数']==last_dt_history)]['单周干预偏差率'].abs().mean()
            curr_avg_huanbiganyu = df_ganyu_kpi[(df_ganyu_kpi['周数']==last_dt_history)]['环比干预偏差率'].abs().mean()

            df_country_turnover = df_country_turnover[df_country_turnover['周数']==last_dt_history]
            历史国内在库周转 = ((((df_country_turnover['当周期初在库'] * df_country_turnover['单价']).sum() + (df_country_turnover['下周期初在库'] * df_country_turnover['单价']).sum()) / 2) / ((df_country_turnover['当周周销']*df_country_turnover['单价'])/7).sum()).round(1)
            df_历史海外周转 = df_历史海外周转[df_历史海外周转['周数']==last_dt_history]
            历史海外在库周转 = (((((df_历史海外周转['历史当周期初在库'] * df_历史海外周转['单价']).sum() + (df_历史海外周转['历史下周期初在库'] * df_历史海外周转['单价']).sum()) / 2) / ((df_历史海外周转['历史当周周销']*df_历史海外周转['单价'])/7).sum())).round(1)
            历史海外在途周转 = ((((df_历史海外周转['历史当周期初在途'] * df_历史海外周转['单价']).sum() + (df_历史海外周转['历史下周期初在途'] * df_历史海外周转['单价']).sum()) / 2) / ((df_历史海外周转['历史当周周销']*df_历史海外周转['单价'])/7).sum()).round(1)
            # 剔除清仓状态、新品FBM状态、新品状态且在库为0的SKU
            # 断货率 = df_历史海外周转[df_历史海外周转['状态']=='断货'].shape[0]/df_历史海外周转.shape[0]
            valid_mask = (
                (df_历史海外周转['状态'] != '清仓') & 
                (df_历史海外周转['状态'] != '新品FBM') & 
                ~((df_历史海外周转['状态'] == '新品') & (df_历史海外周转['历史当周期初在库'] == 0))
            )
            df_valid = df_历史海外周转[valid_mask]

            # 3. 计算断货率（防止分母为0报错，加个判断）
            if df_valid.shape[0] > 0:
                断货率 = (df_valid['状态'] == '断货').sum() / df_valid.shape[0]
            else:
                断货率 = 0.0
            c1, c2, c3, c4,c5,c6,c7,c8,c9,c10 = st.columns(10)
            st.markdown(f"""
                <style>
                /* 1. 定位整个 Metric 容器，让所有子元素在纵向上居中对齐 */
                [data-testid="stMetric"] {{
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    justify-content: center;
                    width: 100%;
                }}

                /* 2. 强制 Label (标题) 居中 */
                [data-testid="stMetricLabel"] {{
                    display: flex;
                    justify-content: center;
                    width: 100%;
                }}

                /* 3. 强制 Value (大数字) 居中并处理变色 */
                [data-testid="stMetricValue"] {{
                    display: flex;
                    justify-content: center;
                    width: 100%;
                }}
                /* 4. 强制 Delta (目标气泡) 居中 */
                [data-testid="stMetricDelta"] {{
                    display: flex;
                    justify-content: center;
                    width: 100%;
                    margin-left: 0 !important;
                }}

                /* 5. 隐藏箭头 */
                [data-testid="stMetricDelta"] svg {{
                    display: none !important;
                }}

                /* 6. 移除气泡内部文字可能的左间距 */
                [data-testid="stMetricDelta"] > div {{
                    margin-left: 0 !important;
                }}
                </style>
                """, unsafe_allow_html=True)
            c1.metric("断货率",f"{断货率:.1%}",delta="目标: 0%",delta_color="green")
            c2.metric("海外在库周转",历史海外在库周转,"目标: P1:45天,P2:30天",delta_color="green",help="(期初在库x单价+期末在库x单价)/2/(周销x单价/7)")
            c3.metric("海外在途周转",历史海外在途周转,"目标: 60天",delta_color="green",help="(期初在途x单价+期末在途x单价)/2/(周销x单价/7)")
            c4.metric("国内在库周转", f"{历史国内在库周转:.1f}", delta=f"目标: 60天", delta_color="green",help="(国内期初在库x单价+国内期末在库x单价)/2/(周销x单价/7)")
            c5.metric("预测偏差率", f"{curr_avg_yuce:.0%}",delta="目标: 35%",delta_color="green")
            c6.metric("预测偏差率(环比)", f"{curr_avg_huanbiyuce:.0%}",delta="目标: 5%",delta_color="green")
            c7.metric("预测偏差率(LT前)", f"{curr_avg_ltyuce:.0%}",delta="目标: 35%",delta_color="green")
            c8.metric("干预SKU占比", f"{ganyu_intervention_rate:.0%}",delta="目标: 15%",delta_color="green")
            c9.metric("干预偏差率", f"{curr_avg_ganyu:.0%}",delta="目标: 30%",delta_color="green")
            c10.metric("干预偏差率(环比)", f"{curr_avg_huanbiganyu:.0%}",delta="目标: 5%",delta_color="green")    
            



# =========================================================
# 5、核心图表函数 (函数名保持不变，日期格式 yyyy-mm-dd)
# =========================================================
def apply_filters(df, filters):
    if df is None: return None
    if filters is None: return df
    df = df.copy()
    if filters.get("sub_market"): df = df[df["子市场"].isin(filters["sub_market"])]
    if filters.get("category"): df = df[df["品类"].isin(filters["category"])]
    if filters.get("mrpsku"): df = df[df["主料mrpsku"].isin(filters["mrpsku"])]
    return df

@st.fragment
def inventorySales_rate_area(df_stock_turnover,df_历史海外周转, curr_filters):
    df = apply_filters(df_stock_turnover, curr_filters)
    last_dt_history = st.session_state.t0_date.strftime("%Yw%V")
    df_历史海外周转 = apply_filters(df_历史海外周转, curr_filters)
    # 未来库存与周转
    weeks = df['周数'].values[0]
    df_历史海外周转=df_历史海外周转[df_历史海外周转["周数"] == weeks]
    # df=df[df["周数"] == last_dt_history]

    market_filter, category_filter, sku_filter,outStock_filter = st.columns(4)
    with market_filter:
        market_list_stock = ["全部市场"] + sorted(df["子市场"].unique().tolist())
        selected_market = st.selectbox("选择要查看的子市场", market_list_stock,key="selectbox_market_stock")
        st.session_state.stock_filter_market=selected_market
    with category_filter:
        category_list_stock = ["全部品类"] + sorted(df["品类"].unique().tolist())
        selected_category = st.selectbox("选择要查看的品类", category_list_stock,key="selectbox_category_stock")
        st.session_state.stock_filter_category=selected_category
    with sku_filter:
        sku_list_stock = ["全部SKU"] + sorted(df["主料mrpsku"].unique().tolist())
        selected_sku = st.selectbox("选择要查看的SKU", sku_list_stock,key="selectbox_sku_stock")
        st.session_state.stock_filter_sku=selected_sku
    with outStock_filter:
        outStock_status_list = ["全部状态","断货","非断货"]
        select_stockStatus = st.selectbox("选择SKU的状态", outStock_status_list,key="selectbox_outStock_stock")
        st.session_state.stock_filter_outStockStatus=select_stockStatus
    
    df_filtered = df.copy()
    if selected_market != "全部市场":
        df_filtered = df_filtered[df_filtered["子市场"] == selected_market]
    if selected_category != "全部品类":
        df_filtered = df_filtered[df_filtered["品类"] == selected_category]
    if selected_sku != "全部SKU":
        df_filtered = df_filtered[df_filtered["主料mrpsku"] == selected_sku]
    if select_stockStatus == "断货":
        df_filtered = df_filtered[df_filtered["状态"] == "断货"]
    elif select_stockStatus == "非断货":
        df_filtered = df_filtered[df_filtered["状态"] != "断货"]

    df_历史海外周转_过滤 = df_历史海外周转.copy()
    if selected_market != "全部市场":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["子市场"] == selected_market]
    if selected_category != "全部品类":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["品类"] == selected_category]
    if selected_sku != "全部SKU":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["主料mrpsku"] == selected_sku]
    if select_stockStatus == "断货":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["状态"] == "断货"]
    elif select_stockStatus == "非断货":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["状态"] != "断货"]
   
    df_temp_future = df_filtered.copy()
    df_temp_future['期初在库金额'] = df_temp_future['当周期初在库']*df_temp_future['单价']
    df_temp_future['期末在库金额'] = df_temp_future['下周期初在库']*df_temp_future['单价']
    df_temp_future['期初在途金额'] = df_temp_future['当周期初在途']*df_temp_future['单价']
    df_temp_future['期末在途金额'] = df_temp_future['下周期初在途']*df_temp_future['单价']
    df_temp_future['当周周销金额'] = np.minimum(df_temp_future['当周期初在库'] + df_temp_future['当周到货'],df_temp_future['当周周销'])*df_temp_future['单价']
    df_temp_future_group = df_temp_future.groupby(['周数'])[['当周期初在库','下周期初在库','当周期初在途',"下周期初在途","单价","期初在库金额","期末在库金额","期初在途金额","期末在途金额","当周到货","当周周销金额","当周周销"]].sum().reset_index().sort_values('周数')
    df_temp_future_group['未来海外在库周转'] = (((df_temp_future_group['期初在库金额'] + df_temp_future_group['期末在库金额'])/2) / (df_temp_future_group['当周周销金额'] / 7)).round(1)
    df_temp_future_group['未来海外在途周转'] = (((df_temp_future_group['期初在途金额'] + df_temp_future_group['期末在途金额'])/2) / (df_temp_future_group['当周周销金额'] / 7)).round(1)

    st.markdown("### 看未来")
    df_temp_future_group=df_temp_future_group[(df_temp_future_group['周数']<='2026w53')]
    fig = make_subplots(
            rows=1, cols=2, 
            shared_yaxes=False,    
            horizontal_spacing=0.08,
            subplot_titles=("海外在库周转", "海外在途周转")
        )

    metrics = ['未来海外在库周转', '未来海外在途周转']
    colors = ['#1f77b4', '#ff7f0e']

    week_values = df_temp_future_group["周数"].values
    week_label = [str(x)[2:] for x in week_values]

    # 2. 循环添加曲线 (row 固定为 1，col 随循环变化)
    for i, col_name in enumerate(metrics):
        col_idx = i + 1
        fig.add_trace(
            go.Scatter(
                x=week_values, 
                y=df_temp_future_group[col_name],
                name=col_name,
                mode='lines+markers+text',
                line=dict(color=colors[i], width=2),
                marker=dict(size=6),
                hovertemplate="周数: %{x}<br>数值: %{y}<extra></extra>",
                textfont=dict(size=12, color="black"),
                textposition="top center",
                text=df_temp_future_group[col_name].values.round(0),
            ),
            row=1, col=col_idx
        )

    for i, annotation in enumerate(fig['layout']['annotations']):
        annotation['font'] = dict(family="Microsoft YaHei", size=20, color="black")
    
    fig.add_hline(y=45,line_dash="dash",line_color="#CC0033", line_width=2,row=1, col=1,annotation_text="P1目标：45天",annotation_position="top right",opacity=0.7,annotation_font=dict(size=10, color="gray"))
    fig.add_hline(y=30,line_dash="dash",line_color="#CC0033", line_width=2,row=1, col=1,annotation_text="P2目标：30天",annotation_position="bottom right",opacity=0.7,annotation_font=dict(size=10, color="gray"))
    fig.add_hline(y=60,line_dash="dash",line_color="#CC0033", line_width=2,row=1, col=2,annotation_text="目标：60天",annotation_position="bottom right" ,opacity=0.7,annotation_font=dict(size=10, color="gray"))
    # 3. 布局优化
    fig.update_layout(
        height=400,
        showlegend=False,
        template="simple_white",
        hovermode="x unified",
        margin=dict(t=60, b=50, l=40, r=40),
        font=dict(family="Microsoft YaHei",size=10),
    )
    fig.update_xaxes(
        ticktext=week_label,
        tickvals=week_values,
        tickfont=dict(size=12, color="gray"),
        title_text="周数",
        showline=True,
        linewidth=1,
        linecolor='lightgray'
    )

    # 5. 美化所有 Y 轴
    fig.update_yaxes(
        showgrid=True, 
        gridcolor='whitesmoke',
        range=[0, 110]
    )

    # 6. 在 Streamlit 中显示
    st.plotly_chart(fig, width='stretch')

    st.markdown("### 海外周转-历史(上周)与目标")
    # 添加筛选框
    在库周转问题筛选,在途周转问题筛选,下载文件按钮=st.columns([2,2,2])
    with 在库周转问题筛选:
        stockTurnover_status_list = ["全部","无问题","预测偏差过大","干预偏差过大","发货问题"]
        select_in_stockStatus = st.selectbox("海外在库周转问题", stockTurnover_status_list)
        # st.session_state.stock_filter_stockTurnover_status=select_in_stockStatus
    with 在途周转问题筛选:
        wayTurnover_status_list = ["全部","无问题","预测偏差过大","干预偏差过大","发货问题"]
        select_in_wayStatus = st.selectbox("海外在途周转问题", wayTurnover_status_list)
        # st.session_state.stock_filter_wayTurnover_status=select_in_wayStatus
    
    if select_in_stockStatus != "全部":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["在库周转问题定义"] == select_in_stockStatus]
    if select_in_wayStatus != "全部":
        df_历史海外周转_过滤 = df_历史海外周转_过滤[df_历史海外周转_过滤["在途周转问题定义"] == select_in_wayStatus]

    with 下载文件按钮:
        st.write(" ")
        st.write(" ")
        def to_excel(df):
            output = io.BytesIO()
            # 使用 xlsxwriter 作为引擎
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
            
            processed_data = output.getvalue()
            return processed_data
        
        excel_data = to_excel(df_历史海外周转_过滤)
        st.download_button(
            label="📥下载 上周海外周转数据",
            data=excel_data,
            file_name=f'上周海外周转数据_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    
    

    st.dataframe(
        df_历史海外周转_过滤[['周数','子市场','主料mrpsku','品类','状态','历史海外在库周转','历史海外在途周转',"海外在库周转目标","海外在途周转目标",'历史当周周销','仿真当周周销','在库周转问题定义','在途周转问题定义']],
        column_config={
            "周数": st.column_config.TextColumn("周数", width=10),
            "子市场": st.column_config.TextColumn("子市场", width=1),
            "品类": st.column_config.TextColumn("品类", width=1),
            "主料mrpsku": st.column_config.TextColumn("MRPSKU", width=50),
            "状态": st.column_config.TextColumn("状态", width=2),
            "历史当周周销": st.column_config.NumberColumn("真实周销", format="%d", width=10, alignment="center"),
            "仿真当周周销": st.column_config.NumberColumn("仿真周销", format="%d", width=10, alignment="center"),
            "历史海外在库周转": st.column_config.NumberColumn("真实海外在库周转", format="%d", width=20, alignment="center"),
            "历史海外在途周转": st.column_config.NumberColumn("真实海外在途周转", format="%d", width=20, alignment="center"),
            "海外在库周转目标": st.column_config.NumberColumn("海外在库周转目标", format="%d", width=10, alignment="center"),
            "海外在途周转目标": st.column_config.NumberColumn("海外在途周转目标", format="%d", width=10, alignment="center"),
            "在库周转问题定义": st.column_config.TextColumn("在库周转问题定义", width=50),
            "在途周转问题定义": st.column_config.TextColumn("在途周转问题定义", width=50),
        },
        hide_index=True,
        width="stretch",
        height=500
    )



    st.markdown("### 未来库存、销量与到货")
    fig = go.Figure()
    df_temp_future_group=df_temp_future_group[(df_temp_future_group['周数']<='2026w53')]

    color_stock = "#1f77b4"
    color_replenish = "#ff7f0e"
    color_sales = "#009966"
    fig.add_trace(go.Scatter(
        x=df_temp_future_group["周数"], 
        y=df_temp_future_group["当周期初在库"],
        mode='lines+text+markers',
        name='库存数量',
        line=dict(color=color_stock, width=3, shape='spline'),
        # text=[f"{row:,.0f}" for row in df_temp_future_group['当周期初在库']],
        # textposition="bottom center",
        # textfont=dict(color=color_stock)
    ))

    fig.add_trace(go.Scatter(
        x=df_temp_future_group["周数"], 
        y=df_temp_future_group["当周到货"],
        mode='lines+text+markers',
        name='到货数量',
        line=dict(color=color_replenish, width=3, shape='spline'),
        # text=[f"{row:,.0f}" for row in df_temp_future_group['当周到货']],
        # textposition="top center",
        # textfont=dict(color=color_replenish)
    ))
    fig.add_trace(go.Scatter(
        x=df_temp_future_group["周数"], 
        y=df_temp_future_group["当周周销"],
        mode='lines+text+markers',
        name='周销数量',
        line=dict(color=color_sales, width=3, shape='spline'),
        # text=[f"{row:,.0f}" for row in df_temp_future_group['当周周销']],
        # textposition="bottom center",
        # textfont=dict(color=color_sales)
    ))
    top_y1 = max(df_temp_future_group["当周期初在库"].max(), df_temp_future_group["当周周销"].max()) * 1.4
    for i, row in df_temp_future_group.iterrows():
        start_y = max(row["当周期初在库"], row["当周周销"])  
        fig.add_shape(
            type="line",
            x0=row["周数"], y0=start_y + 10,
            x1=row["周数"], y1=top_y1 - 20,
            line=dict(color="gray", width=1, dash="dot"),
        )
        fig.add_trace(go.Scatter(
            x=[row["周数"]],
            y=[top_y1],
            mode="markers+text",
            marker=dict(color="white", size=10, line=dict(color="#52c41a", width=2)),
            # text=[f"{row['库销比天数']:.0f}"],
            text=[f"{row['未来海外在库周转']:.0f}"],
            textposition="top center",
            textfont=dict(color="#52c41a", size=14, family="Arial Black"),
            showlegend=i == 0,
            hoverinfo='skip',
            name='未来海外在库周转'
        ))
    new_labels = [str(x)[2:] for x in df_temp_future["周数"].values]
    fig.update_layout(
        plot_bgcolor="white",
        height=500,
        margin=dict(t=80, b=20, l=20, r=20),
        xaxis=dict(
            showgrid=False,
            showline=True,
            linecolor="lightgray",
            ticktext=new_labels,  
            tickvals=df_temp_future["周数"].values,
            tickfont=dict(size=14, color="gray")
        ),
        yaxis=dict(
            showgrid=True,
            showticklabels=True
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
        font=dict(family="Microsoft YaHei",size=10),
    )
    st.plotly_chart(fig, width='stretch')

    st.markdown("### 海外在库在途周转区间")
    df_future_single = df_filtered[(df_filtered['周数']<='2026w53')].copy()
    df_future_single['期初在库金额'] = df_future_single['当周期初在库']*df_future_single['单价']
    df_future_single['期末在库金额'] = df_future_single['下周期初在库']*df_future_single['单价']
    df_future_single['期初在途金额'] = df_future_single['当周期初在途']*df_future_single['单价']
    df_future_single['期末在途金额'] = df_future_single['下周期初在途']*df_future_single['单价']
    # df_future_single['当周周销金额'] = df_future_single['当周周销']*df_future_single['单价']
    df_future_single['当周周销金额'] = np.minimum(df_future_single['当周期初在库'] + df_future_single['当周到货'],df_future_single['当周周销'])*df_future_single['单价']
    df_future_single_oneweek = df_future_single.groupby(['子市场','品类','主料mrpsku'])[['当周期初在库','下周期初在库','当周期初在途',"下周期初在途","单价","期初在库金额","期末在库金额","期初在途金额","期末在途金额","当周到货","当周周销金额","当周周销"]].sum().reset_index()
    df_future_single_oneweek['未来海外在库周转'] = (((df_future_single_oneweek['期初在库金额'] + df_future_single_oneweek['期末在库金额'])/2) / (df_future_single_oneweek['当周周销金额'] / 7)).round(1)
    df_future_single_oneweek['未来海外在途周转'] = (((df_future_single_oneweek['期初在途金额'] + df_future_single_oneweek['期末在途金额'])/2) / (df_future_single_oneweek['当周周销金额'] / 7)).round(1)
    # df_future_single_oneweek = df_future_single[df_future_single['周数']==min(df_future_single['周数'])]
    # df_future_single_oneweek=df_future_single.groupby(['子市场','品类','主料mrpsku'])[['未来海外在库周转','未来海外在途周转']].mean().reset_index()
    周转SKU总数 = len(df_future_single_oneweek)
    df_future_single_oneweek['未来海外在库周转区间'] = pd.cut(
        df_future_single_oneweek['未来海外在库周转'], 
        bins=[0, 15, 30, 45, float('inf')], 
        labels=['[0,15)', '[15,30)', '[30,45)', '[45,+inf)']
    )
    df_future_single_oneweek['未来海外在途周转区间'] = pd.cut(
        df_future_single_oneweek['未来海外在途周转'], 
        bins=[0, 30,45, 60, float('inf')], 
        labels=['[0,30)', '[30,45)', '[45,60)', '[60,+inf)']
    )
    
    df_future_single_oneweek['在库下一步动作']=df_future_single_oneweek['未来海外在库周转区间'].map({
        '[0,15)': '快速补SS',
        '[15,30)': '补SS',
        '[30,45)': '稳定补SS',
        '[45,+inf)': '控制补货'
    })
    df_future_single_oneweek['在途下一步动作']=df_future_single_oneweek['未来海外在途周转区间'].map({
        '[0,30)': '快速补SS',
        '[30,45)': '补SS',
        '[45,60)': '稳定补SS',
        '[60,+inf)': '控制补货'
    })
    在库下一步动作筛选,在途下一步动作筛选,下一步动作下载按钮 = st.columns(3)
    with 在库下一步动作筛选:
        在库下一步动作列表 = ["全部","快速补SS","补SS","稳定补SS","控制补货"]
        选择在库下一步动作 = st.selectbox("在库动作", 在库下一步动作列表)
    with 在途下一步动作筛选:
        在途下一步动作列表 =  ["全部","快速补SS","补SS","稳定补SS","控制补货"]
        选择在途下一步动作 = st.selectbox("在途动作", 在途下一步动作列表)

    if 选择在库下一步动作 != '全部':
        df_future_single_oneweek = df_future_single_oneweek[df_future_single_oneweek['在库下一步动作']==选择在库下一步动作]
    if 选择在途下一步动作 != '全部':
        df_future_single_oneweek = df_future_single_oneweek[df_future_single_oneweek['在途下一步动作']==选择在途下一步动作]


    with 下一步动作下载按钮:
        st.write(" ")
        st.write(" ")
        def to_excel(df):
            output = io.BytesIO()
            # 使用 xlsxwriter 作为引擎
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
            
            processed_data = output.getvalue()
            return processed_data
        
        excel_data = to_excel(df_future_single_oneweek)
        st.download_button(
            label="📥下载 海外在库在途下一步动作",
            data=excel_data,
            file_name=f'海外在库在途下一步动作_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    stock_col1,stock_col2 = st.columns([1,3])
    with stock_col1:
        st.markdown("#### SKU个数汇总")
        stock_counts = df_future_single_oneweek['未来海外在库周转区间'].value_counts().sort_index().reset_index()
        stock_counts.columns = ['区间', '数量']
        stock_counts['维度'] = '未来海外在库'

        transit_counts = df_future_single_oneweek['未来海外在途周转区间'].value_counts().sort_index().reset_index()
        transit_counts.columns = ['区间', '数量']
        transit_counts['维度'] = '未来海外在途' 
        plot_df = pd.concat([stock_counts, transit_counts]).dropna(subset=['区间'])
        
        stock_color_map = {
            '[0,15)': '#f8c471', '[15,30)': '#85C1E9', '[30,45)': '#82E0AA', '[45,+inf)': '#F79C75'
        }
        onway_color_map={
            '[0,30)': '#f8c471', '[30,45)': '#85C1E9', '[45,60)': '#82E0AA', '[60,+inf)': '#F79C75'
        }
        # 将颜色映射到 DataFrame 中
        def get_color(row):
            if row['维度'] == '未来海外在库':
                return stock_color_map.get(row['区间'])
            else:
                return onway_color_map.get(row['区间'])
        plot_df['颜色'] = plot_df.apply(get_color, axis=1)
        # st.write(plot_df)
        fig_turnover=go.Figure()
        fig_turnover.add_trace(go.Bar(
            name='SKU总数',
            x=['SKU总数'],
            y=[周转SKU总数],
            text=[周转SKU总数],
            textposition='outside',
            marker_color="gray",
            textfont=dict(size=14,family="Microsoft YaHei", color="black")
        ))
        dimensions_order = ['SKU总数', '未来海外在库', '未来海外在途']
        interval_rank = {
            "SKU总数": 0,
            "[0,15)": 10, "[15,30)": 20, "[30,45)": 30, "[45,+inf)": 40, # 在库
            "[0,30)": 11, "[30,45)": 21, "[45,60)": 31, "[60,+inf)": 41  # 在途 (微调权重确保排序)
        }
        plot_df['rank'] = plot_df['区间'].map(interval_rank)
        plot_df = plot_df.sort_values(by=['维度', 'rank'], ascending=True)
        SMALL_VALUE_THRESHOLD = 100
        for dim in dimensions_order:
            dim_data = plot_df[plot_df['维度'] == dim]
            cumulative_y = 0 
            for _, row in dim_data.iterrows():
                interval = row['区间']
                value = row['数量']
                if dim == '未来海外在库':
                    current_color = stock_color_map.get(interval)
                elif dim == '未来海外在途':
                    current_color = onway_color_map.get(interval)
                else:
                    current_color = 'gray'
                show_text = value >= SMALL_VALUE_THRESHOLD
                fig_turnover.add_trace(go.Bar(
                    x=[dim],
                    y=[value],
                    width=0.4,
                    name=f"{dim}-{interval}",
                    marker_color=current_color,
                    text=[value] if show_text else [""],
                    textposition='inside' if show_text else 'none',
                    legendgroup=dim,
                    hovertemplate=f"维度: {dim}<br>区间: {interval}<br>数量: %{{y}}<extra></extra>",
                    textfont=dict(
                        size=14,
                        family="Microsoft YaHei",
                        color="black"
                    ),
                ))
                if value < SMALL_VALUE_THRESHOLD and value > 0:
                    fig_turnover.add_annotation(
                        x=dim,
                        y = cumulative_y + value / 2,
                        text=str(value),
                        showarrow=True,
                        arrowhead=2,
                        arrowsize=1,
                        arrowwidth=1.5,
                        arrowcolor="black",
                        ax=50,   # 水平方向偏移
                        ay=-10,  # 垂直方向偏移
                        font=dict(
                            size=12,
                            color="black"
                        ),
                        bgcolor="rgba(255,255,255,0.8)"
                    )
                cumulative_y += value

        # 6. 配置布局
        fig_turnover.update_xaxes(
            tickfont=dict(size=14, color="black")
        )
        fig_turnover.update_layout(
            barmode='relative',
            bargap=0.5,
            xaxis=dict(
                categoryorder='array', 
                categoryarray=dimensions_order,
                color='gray',
            ),
            legend=dict(
                orientation="h",
                traceorder="normal",
                groupclick="toggleitem" 
            ),
            legend_tracegroupgap=40,
            legend_title="周转区间",
            margin=dict(t=15),
        )
        st.plotly_chart(fig_turnover, width='stretch',height=500)

    with stock_col2:
        st.markdown("#### SKU明细")
        st.dataframe(
            df_future_single_oneweek[['子市场','品类','主料mrpsku','未来海外在库周转','未来海外在途周转','未来海外在库周转区间','未来海外在途周转区间','在库下一步动作','在途下一步动作']],
            column_config={
                "子市场": st.column_config.TextColumn("子市场", width=20),
                "品类": st.column_config.TextColumn("品类", width=20),
                "主料mrpsku": st.column_config.TextColumn("MRPSKU", width=30),
                "未来海外在库周转": st.column_config.NumberColumn("海外在库周转(目标:45天)", format="%d", width=20, alignment="center"),
                "未来海外在途周转": st.column_config.NumberColumn("海外在途周转(目标:60天)", format="%d", width=20, alignment="center"),
                "未来海外在库周转区间": st.column_config.TextColumn("海外在库周转区间", width=20),
                "未来海外在途周转区间": st.column_config.TextColumn("海外在途周转区间", width=20),
                "在库下一步动作": st.column_config.TextColumn("在库下一步动作", width=20),
                "在途下一步动作": st.column_config.TextColumn("在途下一步动作", width=20)
            },
            hide_index=True,
            width="stretch",
            height=500
        )

  
# 预测指标区域
@st.fragment
def predictSales_rate_area(df_yuce, curr_filters):
    df = apply_filters(df_yuce, curr_filters)
    if df is None or df.empty:
        st.warning("无数据")
        return

    # st.subheader("📊 预测准确度监控看板")
    # --- 1. 顶部 KPI 总览 ---
    
    market_filter, category_filter, sku_filter,outStock_filter,week_filter = st.columns(5)
    with market_filter:
        market_list = ["全部市场"] + sorted(df["子市场"].unique().tolist())
        selected_market = st.selectbox("选择要查看的子市场", market_list,key="selectbox_market_yuce")
        st.session_state.yuce_filter_market=selected_market
    with category_filter:
        category_list = ["全部品类"] + sorted(df["品类"].unique().tolist())
        selected_category = st.selectbox("选择要查看的品类", category_list,key="selectbox_category_yuce")
        st.session_state.yuce_filter_category=selected_category
    with sku_filter:
        SKU_list = ["全部SKU"] + sorted(df["主料mrpsku"].unique().tolist())
        selected_sku = st.selectbox("选择要查看的SKU", SKU_list,key="selectbox_sku_yuce")
        st.session_state.yuce_filter_sku=selected_sku
    with outStock_filter:
        outStock_status_list = ["全部状态","断货","非断货"]
        select_stockStatus = st.selectbox("选择SKU的状态", outStock_status_list,key="selectbox_outStock_yuce")
        st.session_state.yuce_filter_outStockStatus=select_stockStatus
    with week_filter:
        # default_monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
        if st.session_state.t0_date is not None:
            default_monday=st.session_state.t0_date.strftime("%Y-%m-%d")
        else:
            default_monday = (today - timedelta(days=today.weekday()) - timedelta(days=7)).strftime("%Y-%m-%d")

        select_week = st.date_input("选择要查看的周数", value=default_monday,key="date_input_week_yuce")
        select_week = select_week.strftime("%Yw%V")
        st.session_state.yuce_filter_week=select_week


    df_filtered = df.copy()
    if selected_market != "全部市场":
        df_filtered = df_filtered[df_filtered["子市场"] == selected_market]

    if selected_category != "全部品类":
        df_filtered = df_filtered[df_filtered["品类"] == selected_category]

    if selected_sku != "全部SKU":
        df_filtered = df_filtered[df_filtered["主料mrpsku"] == selected_sku]

    if select_stockStatus == "断货":
        df_filtered = df_filtered[df_filtered["状态"] == "断货"]
    elif select_stockStatus == "非断货":
        df_filtered = df_filtered[df_filtered["状态"] != "断货"]

    # df_filtered_noOutStock = df_filtered[df_filtered["状态"] != "断货"]
    单周预测偏差 = df_filtered[df_filtered["周数"] == select_week]["单周预测偏差率"].abs().mean()
    环比预测偏差 = df_filtered[df_filtered["周数"] == select_week]["环比预测偏差率"].abs().mean()
    m_col1, m_col2 = st.columns(2)
    m_col1.metric("预测偏差率", f"{单周预测偏差:.0%}", delta="目标:35%")
    m_col2.metric("预测偏差率(环比)", f"{环比预测偏差:.0%}", delta="目标:5%")
    col_left, col_right = st.columns(2)
    stage_cols = ["单周预测偏差率", "环比预测偏差率"]
    df_filtered[stage_cols] = df_filtered[stage_cols].abs()
    df_m = df_filtered[df_filtered["周数"] == select_week].groupby("子市场").agg({
        "单周预测偏差率": "mean",
        "环比预测偏差率": "mean",
        "主料mrpsku": "count"
    }).reset_index()
    df_m = df_m.rename(columns={"主料mrpsku": "SKU数量"})
    df_m = df_m.sort_values(by="SKU数量", ascending=True)

    with col_left:
        st.markdown("#### 历史预测偏差趋势")
        df_filtered_week = df_filtered.groupby("周数").agg({
            "单周预测偏差率": "mean",
            "环比预测偏差率": "mean"
        }).reset_index()
        fig_历史预测曲线=go.Figure()
        fig_历史预测曲线.add_trace(
            go.Scatter(
                x=df_filtered_week["周数"],
                y=df_filtered_week["单周预测偏差率"],
                mode='lines+markers+text',
                line=dict(color='#1f77b4', width=2),
                marker=dict(
                    color="#1f77b4", 
                    size=20, 
                    line=dict(width=5, color='white')
                ),
                text=df_filtered_week["单周预测偏差率"].apply(lambda x: f"{x:.0%}"),
                textposition="top left",
                name="单周偏差",
                textfont=dict(size=14, color="black")
            )
        )
        fig_历史预测曲线.add_trace(
            go.Scatter(
                x=df_filtered_week["周数"],
                y=df_filtered_week["环比预测偏差率"],
                mode='lines+markers+text',
                line=dict(color='#ff7f0e', width=2),
                marker=dict(
                    color="#ff7f0e", 
                    size=20, 
                    line=dict(width=5, color='white')
                ),
                text=df_filtered_week["环比预测偏差率"].apply(lambda x: f"{x:.0%}"),
                textposition="bottom left",
                name="环比偏差",
                yaxis="y2",
                textfont=dict(size=14, color="black")
            )
        )
        fig_历史预测曲线.add_hline(
            y=0.35,
            line_dash="dash",
            line_color="#1f77b4",  
            annotation_text="预测偏差: 35%",
            annotation_position="bottom right",
            opacity=0.7,
            annotation_font=dict(size=12, color="gray")
        )
        fig_历史预测曲线.add_shape(
            type="line",
            xref="paper", yref="y2",  # x轴跨越整个画布，y轴绑定到右轴 y2
            x0=0, y0=0.05,
            x1=1, y1=0.05,
            line=dict(color="#ff7f0e", width=2, dash="dash"),
            opacity=0.7
            # label=dict(
            #     text="环比预测偏差: 35%",           
            #     font=dict(color="gray", size=12),
            #     textposition="bottom right",  
            #     yanchor="bottom"               
            # )
        )
        fig_历史预测曲线.update_layout(
            # 左侧 Y 轴配置
            yaxis=dict(
                title="单周预测偏差率",
                range=[0, 1.0],  
                tickformat=".0%", 
                side="left"
            ),
            # 右侧 Y 轴配置
            yaxis2=dict(
                title="环比预测偏差率",
                range=[0, 0.2],  
                tickformat=".0%", 
                overlaying="y",   
                side="right"
            ),
            # 可选：优化整体布局，防止右侧 Y 轴标题被截断
            margin=dict(r=80),
            legend=dict(
                orientation="h",          # 水平排列
                yanchor="bottom",         # 垂直方向以底部为锚点
                y=1.02,                   # 放在绘图区顶部稍微偏上的位置
                xanchor="center",         # 水平方向以中心为锚点
                x=0.5,                    # 放在水平方向 50% 的位置（居中）
                bgcolor="rgba(255,255,255,0.8)" # 可选：设置半透明背景，防止遮挡标题或边框
            ),
            font=dict(family="Microsoft YaHei",size=10)
        )

        st.plotly_chart(fig_历史预测曲线,width="stretch",height=550)


    with col_right:
        TARGET_MAX = 0.35
        st.markdown("#### 各市场预测偏差全貌 (目标区间:35%)")

        # --- 2. 创建布局 ---
        fig = make_subplots(
            rows=1, cols=2, 
            shared_yaxes=True, 
            horizontal_spacing=0.03, 
            column_widths=[0.7, 0.3]
        )
        colors_bar = [
            "#ff9d4f" if (x > TARGET_MAX) else "#5da9c4" 
            for x in df_m["单周预测偏差率"]
        ]

        fig.add_trace(
            go.Bar(
                y=df_m["子市场"],
                x=df_m["单周预测偏差率"],
                orientation='h',
                marker_color=colors_bar,
                text=df_m["单周预测偏差率"].apply(lambda x: f"{x:.1%}"),
                textposition='outside',
                cliponaxis=False,
                name="单周偏差",
                textfont=dict(size=14, color="black")
            ),
            row=1, col=1
        )

        for line_x in [TARGET_MAX]:
            fig.add_vline(
                x=line_x, 
                line_dash="dash", 
                line_color="red", 
                line_width=2,
                row=1, col=1,
                # 不透明度
                opacity=0.8
            )


        # --- 5. 右侧：环比预测偏差率 (折线图) ---
        colors_line = ["#5da9c4" if abs(x) <= 0.05 else "#f5222d" for x in df_m["环比预测偏差率"]]

        fig.add_trace(
            go.Scatter(
                y=df_m["子市场"],
                x=df_m["环比预测偏差率"],
                mode='lines+markers+text',
                line=dict(color='#bfbfbf', width=3),
                marker=dict(
                    color=colors_line, 
                    size=20, 
                    line=dict(width=5, color='white')
                ),
                text=df_m["环比预测偏差率"].apply(lambda x: f"{x:.0%}"),
                textposition="middle right",
                name="环比趋势",
                textfont=dict(size=14, color="black")
            ),
            row=1, col=2
        )
        for line_x in [0.05, -0.05]:
            fig.add_vline(
                x=line_x, 
                line_dash="dash", 
                line_color="red", 
                line_width=2,
                row=1, col=2,
                # 不透明度
                opacity=0.8
            )

        # --- 6. 布局精修 ---
        # 处理 UK 等极端值：确保坐标轴能盖住 ±30%
        max_val = max(df_m["单周预测偏差率"].max(), 0.5)

        fig.update_layout(
            template="simple_white",
            showlegend=False,
            height=len(df_m) * 45 + 120,
            margin=dict(l=10, r=60, t=20, b=40),
            xaxis=dict(
                title="单周偏差(35%)",
                tickformat=".0%",
                range=[0, max_val + 0.3], # 留出文本空间
                zeroline=True, zerolinecolor="#8c8c8c"
            ),
            xaxis2=dict(
                title="环比偏差(5%)",
                tickformat=".0%",
                range=[0, 0.5], # 环比范围固定，方便观察斜率
                showgrid=False,
                zeroline=True
            ),
            font=dict(family="Microsoft YaHei",size=10),
        )
        

        # 样式细节：隐藏右图Y轴刻度，统一字体
        fig.update_yaxes(showgrid=False, row=1, col=1)
        fig.update_yaxes(showticklabels=False, row=1, col=2)
        fig.update_yaxes(tickfont=dict(size=13), row=1, col=1)

        st.plotly_chart(fig, width='stretch')
            
    df_cat = df_filtered[df_filtered["周数"] == select_week].groupby("品类")[stage_cols].mean().reset_index()
    df_cat = df_filtered.groupby("品类").agg({
        "单周预测偏差率": "mean",
        "环比预测偏差率": "mean",
        "主料mrpsku": "nunique" 
    }).reset_index()
    df_cat.columns = ["品类", "单周预测偏差率", "环比预测偏差率", "SKU个数"]
    def get_color(row):
        if row["单周预测偏差率"] > 0.27:
            return "预测过高"
        if row["单周预测偏差率"] < -0.27:
            return "预测过低"
        return "正常"
    
    df_cat["偏差情况"] = df_cat.apply(get_color, axis=1) 
    detail_cat,detail_sku = st.columns([2,3])
    df_cat['单周预测偏差率'] = round(df_cat['单周预测偏差率']*100, 1)
    df_cat['环比预测偏差率'] = round(df_cat['环比预测偏差率']*100, 1)
    with detail_cat:
        markdown_text,download_button = st.columns([4,2])
        markdown_text.markdown(
            f"""
            #### 预测情况下钻-<span style='color: #ff4b4b;'>{st.session_state.yuce_filter_market}</span> 品类明细表
            """, 
            unsafe_allow_html=True
        )
        def to_excel(df):
            output = io.BytesIO()
            # 使用 xlsxwriter 作为引擎
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
            
            processed_data = output.getvalue()
            return processed_data
        excel_data = to_excel(df_cat)
        download_button.download_button(
            label="📥下载 品类明细表",
            data=excel_data,
            file_name=f'品类明细表_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        def color_deviation(val,target):
            color = '#cf1322' if abs(val) > target else '#389e0d'
            return f'color: {color}; font-weight: bold' if abs(val) > target else f'color: {color}'
        df_cat["周数"] = select_week
        styled_df = df_cat[['周数','品类', '单周预测偏差率', '环比预测偏差率', 'SKU个数', '偏差情况']].sort_values("单周预测偏差率", ascending=False).style.map(
            color_deviation, subset=['单周预测偏差率'],target=35
        )
        styled_df = styled_df.map(
            color_deviation, subset=['环比预测偏差率'],target=5
        )

        
        st.dataframe(
            styled_df,
            column_config={
                "单周预测偏差率": st.column_config.NumberColumn(
                    "单周预测偏差率(目标值:35%)",
                    help="预测值与实际值的偏移程度",
                    format="%.1f%%",
                    width=80,
                    alignment="center"
                ),
                "环比预测偏差率": st.column_config.NumberColumn(
                    "环比预测偏差率(目标值:5%)",
                    help="本周对比上周偏差的变化",
                    format="%.1f%%",
                    width=80,
                    alignment="center"
                ),
                "SKU个数": st.column_config.NumberColumn(
                    "SKU个数",
                    format="%d",
                    width=50,
                    alignment="center"
                ),
                "偏差情况": st.column_config.TextColumn("偏差情况", width=80),
        },
         width="stretch", height=300, hide_index=True)
    with detail_sku:
        markdown_text,download_button = st.columns([3,1])
        df_detail = df_filtered[df_filtered["周数"] == select_week][['周数','子市场', "channel_name",'品类', "主料mrpsku","状态", "当周实际值", "当周预测值" ,"单周预测偏差率", "环比预测偏差率"]].copy()
        df_detail=df_detail.sort_values("单周预测偏差率", ascending=False)
        df_detail['单周预测偏差率'] = round(df_detail['单周预测偏差率']*100, 1)
        df_detail['环比预测偏差率'] = round(df_detail['环比预测偏差率']*100, 1)

        markdown_text.markdown(
            f"""
            #### 预测情况下钻-<span style='color: #ff4b4b;'>{st.session_state.yuce_filter_market}_{st.session_state.yuce_filter_category}</span> MRPSKU明细表
            """, 
            unsafe_allow_html=True
        )
        def to_excel(df):
            output = io.BytesIO()
            # 使用 xlsxwriter 作为引擎
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
            
            processed_data = output.getvalue()
            return processed_data
        excel_data = to_excel(df_detail)
        download_button.download_button(
            label="📥下载 MRPSKU明细表",
            data=excel_data,
            file_name=f'MRPSKU明细表_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        def color_deviation(val,target):
            color = '#cf1322' if abs(val) > target else '#389e0d'
            return f'color: {color}; font-weight: bold' if abs(val) > target else f'color: {color}'
        styled_df = df_detail[['周数','子市场', "channel_name",'品类', "主料mrpsku","状态", "当周实际值", "当周预测值" ,"单周预测偏差率", "环比预测偏差率"]].sort_values("单周预测偏差率", ascending=False).style.map(
            color_deviation, subset=['单周预测偏差率'],target=35
        )
        styled_df = styled_df.map(
            color_deviation, subset=['环比预测偏差率'],target=5
        )
        st.dataframe(
            styled_df,
            column_config={
                "周数": st.column_config.TextColumn("周数", width=20),
                "子市场": st.column_config.TextColumn("子市场", width=20),
                "channel_name": st.column_config.TextColumn("渠道", width=20),
                "品类": st.column_config.TextColumn("品类", width=20),
                "主料mrpsku": st.column_config.TextColumn("MRPSKU", width=30),
                "状态": st.column_config.TextColumn("状态", width=20),
                "当周实际值": st.column_config.NumberColumn("实际周销", format="%d", width=20, alignment="center"),
                "当周预测值": st.column_config.NumberColumn("预测周销", format="%d", width=20, alignment="center"),
                # "上周预测值": st.column_config.NumberColumn("上周预测周销", format="%d", width=20, alignment="center"),
                "单周预测偏差率": st.column_config.NumberColumn(
                    "单周预测偏差率(目标值:35%)",
                    format="%.1f%%", width=50, alignment="center" 
                ),
                "环比预测偏差率": st.column_config.NumberColumn(
                    "环比预测偏差率(目标值:5%)",
                    format="%.1f%%", width=50, alignment="center" 
                )
            },
            hide_index=True,
            width="stretch",
            height=300
        )
# 干预指标区域
@st.fragment
def ganyuSales_rate_area(df_ganyu, curr_filters):
    df = apply_filters(df_ganyu, curr_filters)
    if df is None or df.empty:
        st.warning("无数据")
        return
    
    # df=df[df['是否有干预'] == '是']
    # st.subheader("📊 预测准确度监控看板")
    market_filter, category_filter, sku_filter,outStock_filter,week_filter = st.columns(5)
    with market_filter:
        market_list = ["全部市场"] + sorted(df["子市场"].unique().tolist())
        selected_market = st.selectbox("选择要查看的子市场", market_list,key="selectbox_market_ganyu")
        st.session_state.ganyu_filter_market=selected_market
    with category_filter:
        category_list = ["全部品类"] + sorted(df["品类"].unique().tolist())
        selected_category = st.selectbox("选择要查看的品类", category_list,key="selectbox_category_ganyu")
        st.session_state.ganyu_filter_category=selected_category
    with sku_filter:
        SKU_list = ["全部SKU"] + sorted(df["主料mrpsku"].unique().tolist())
        selected_sku = st.selectbox("选择要查看的SKU", SKU_list,key="selectbox_sku_ganyu")
        st.session_state.ganyu_filter_sku=selected_sku
    with outStock_filter:
        outStock_status_list = ["全部状态","断货","非断货"]
        select_stockStatus = st.selectbox("选择SKU的状态", outStock_status_list,key="selectbox_outStock_ganyu")
        st.session_state.ganyu_filter_outStockStatus=select_stockStatus
    with week_filter:
        # default_monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
        if st.session_state.t0_date is not None:
            default_monday=st.session_state.t0_date.strftime("%Y-%m-%d")
        else:
            default_monday = (today - timedelta(days=today.weekday()) - timedelta(days=7)).strftime("%Y-%m-%d")
        select_week = st.date_input("选择要查看的周数", value=default_monday,key="date_input_week_ganyu")
        select_week = select_week.strftime("%Yw%V")
        st.session_state.ganyu_filter_week=select_week

    df_filtered = df.copy()
    if selected_market != "全部市场":
        df_filtered = df_filtered[df_filtered["子市场"] == selected_market]

    if selected_category != "全部品类":
        df_filtered = df_filtered[df_filtered["品类"] == selected_category]

    if selected_sku != "全部SKU":
        df_filtered = df_filtered[df_filtered["主料mrpsku"] == selected_sku]

    if select_stockStatus == "断货":
        df_filtered = df_filtered[df_filtered["状态"] == "断货"]
    elif select_stockStatus == "非断货":
        df_filtered = df_filtered[df_filtered["状态"] != "断货"]
    
    df_filtered_weeks = df_filtered.copy()
    df_filtered_weeks_yes = df_filtered_weeks[df_filtered_weeks["是否有干预"] == "是"].copy()
    df_filtered = df_filtered[df_filtered["周数"] == select_week]
    df_filtered_yes = df_filtered[df_filtered["是否有干预"] == "是"]
    avg_bias = df_filtered_yes["单周干预偏差率"].abs().mean()
    avg_bias_ratio = df_filtered_yes["环比干预偏差率"].abs().mean()
    ganyu_intervention_rate = len(df_filtered_yes) / len(df_filtered)
    ganyu_effect_intervention_rate = len(df_filtered_yes[df_filtered_yes["是否应该干预"] == "是"]) / len(df_filtered_yes)
    m_col1, m_col2, m_col3,m_col4 = st.columns([1,1,1,1])
    # ganyu_intervention_rate = df_ganyu_bi["有干预样本数"].sum() / df_ganyu_bi["总样本数"].sum()
    m_col1.metric("干预SKU个数占比", f"{ganyu_intervention_rate:.0%}", delta="目标:15%")
    m_col2.metric("有效干预SKU个数占比", f"{ganyu_effect_intervention_rate:.0%}",delta="有效干预规则:±30%")
    m_col3.metric("干预偏差率", f"{avg_bias:.0%}" , delta="目标:30%")
    m_col4.metric("干预偏差率(环比)", f"{avg_bias_ratio:.0%}", delta="目标:5%")

    col_left, col_right = st.columns([2,3])
    # 市场维度聚合
    stage_cols = ["单周干预偏差率", "环比干预偏差率"]
    df_market = df_filtered.groupby("子市场")[stage_cols].mean().reset_index()
    df_plot = df_market.melt(id_vars="子市场", value_vars=stage_cols, 
                            var_name="指标类型", value_name="偏差率")
    # if st.session_state.ganyu_filter_market != "全部市场":
    #     df_ganyu_bi=df_ganyu_bi[df_ganyu_bi["子市场"] == st.session_state.ganyu_filter_market]
    # df_inter = df_ganyu_bi.sort_values("有干预样本数", ascending=False)
    df_filtered_weeks['是否有干预'] = df_filtered_weeks['是否有干预'].map({'是': True, '否': False})
    df_filtered_weeks['是否应该干预'] = df_filtered_weeks['是否应该干预'].map({'是': True, '否': False})
    df_filtered['是否有干预'] = df_filtered['是否有干预'].map({'是': True, '否': False})
    df_filtered['是否应该干预'] = df_filtered['是否应该干预'].map({'是': True, '否': False})
    df_inter = df_filtered_weeks.groupby('周数').agg(
        总样本数=('周数', 'size'),
        有干预样本数=('是否有干预', 'sum'),
        有效干预数=('是否应该干预', 'sum')
    ).reset_index()
    df_inter["不应干预样本数"] = df_inter["有干预样本数"] - df_inter["有效干预数"]
    with col_left:
        st.markdown("#### 历史干预SKU个数占比趋势")
        df_inter_week = df_inter.groupby("周数").agg({
            "总样本数": "sum",
            "有干预样本数": "sum",
        }).reset_index()
        df_inter_week['干预SKU个数占比'] = df_inter_week["有干预样本数"] / df_inter_week["总样本数"]
        fig_历史干预SKU曲线=go.Figure()
        fig_历史干预SKU曲线.add_trace(
            go.Scatter(
                x=df_inter_week["周数"],
                y=df_inter_week["干预SKU个数占比"],
                mode='lines+markers+text',
                line=dict(color='#1f77b4', width=2),
                marker=dict(
                    color="#1f77b4", 
                    size=20, 
                    line=dict(width=5, color='white')
                ),
                text=df_inter_week["干预SKU个数占比"].apply(lambda x: f"{x:.0%}"),
                textposition="top left",
                name="干预SKU个数占比",
                textfont=dict(size=14, color="black")
            )
        )
        fig_历史干预SKU曲线.add_hline(
            y=0.15,
            line_dash="dash",
            line_color="#1f77b4",  
            annotation_text="干预SKU个数占比: 15%",
            annotation_position="bottom right",
            opacity=0.7
        )
        fig_历史干预SKU曲线.update_layout(
            # 左侧 Y 轴配置
            yaxis=dict(
                # title="单周干预偏差率",
                range=[0, 1.0],  
                tickformat=".0%", 
                side="left"
            ),
            margin=dict(r=80),
            legend=dict(
                orientation="h",          # 水平排列
                yanchor="bottom",         # 垂直方向以底部为锚点
                y=1.02,                   # 放在绘图区顶部稍微偏上的位置
                xanchor="center",         # 水平方向以中心为锚点
                x=0.5,                    # 放在水平方向 50% 的位置（居中）
                bgcolor="rgba(255,255,255,0.8)" # 可选：设置半透明背景，防止遮挡标题或边框
            ),
            font=dict(family="Microsoft YaHei",size=12)
        )

        st.plotly_chart(fig_历史干预SKU曲线,width="stretch",height=500)


    with col_right:
        st.markdown("#### 干预SKU个数占比")

        df_inter = df_filtered.groupby('子市场').agg(
            总样本数=('子市场', 'size'),
            有干预样本数=('是否有干预', 'sum'),
            有效干预数=('是否应该干预', 'sum')
        ).reset_index()
        df_inter["不应干预样本数"] = df_inter["有干预样本数"] - df_inter["有效干预数"]
        fig = make_subplots(
            rows=1, cols=2,
            column_widths=[0.35, 0.65], # 左侧占比稍小
            horizontal_spacing=0.12,
            subplot_titles=("市场干预SKU占比分布", "有效干预VS无效干预"),
            specs=[[{"type": "pie"}, {"type": "xy"}]]
        )
        labels = df_inter["子市场"].tolist()
        v_total = df_inter["有干预样本数"].tolist()
        v_effective = df_inter["有效干预数"].tolist()
        # --- 3. 左侧：精美的圆环图 ---
        business_colors = ['#1f77b4', '#4e79a7', '#76b7b2', '#54a24b', '#e15759', '#f28e2b', '#bab0ac', '#86bcb6', '#dbedff']
        fig.add_trace(go.Pie(
            labels=labels,
            values=v_total,
            hole=0.6,
            textinfo='label+percent', 
            textposition='outside',
            automargin=True,
            showlegend=False, 
            marker=dict(
                colors=business_colors, 
                line=dict(color='#FFFFFF', width=2),
            ),
            # marker=dict(line=dict(color='#FFFFFF', width=2)),
            name="市场占比",
            opacity=0.7,
            textfont=dict(size=14, color="black")
        ), row=1, col=1)

        # --- 4. 右侧：簇状柱状图 ---
        # 柱子1：有干预样本数（背景色调）
        fig.add_trace(go.Bar(
            x=labels,
            y=v_total,
            name='有干预样本数',
            marker_color='rgba(55, 128, 191, 0.6)', # 浅蓝色
            offsetgroup=1,
            text=v_total,
            textposition='outside',
            outsidetextfont=dict(size=14, color="black")
        ), row=1, col=2)

        # 柱子2：有效干预数（深色突出）
        fig.add_trace(go.Bar(
            x=labels,
            y=v_effective,
            name='有效干预数',
            marker_color='rgba(50, 171, 96, 1.0)', # 翠绿色
            offsetgroup=2,
            text=v_effective,
            textposition='outside',
            outsidetextfont=dict(size=14, color="black")
        ), row=1, col=2)

        top_y1 = max(v_total) + 100
        for i, row in df_inter.iterrows():
            start_y = max(row["有干预样本数"], row["有效干预数"])  
            fig.add_shape(
                type="line",
                x0=row["子市场"], y0=start_y + 10,
                x1=row["子市场"], y1=top_y1 - 20,
                line=dict(color="gray", width=1, dash="dot"),
            )
            
            fig.add_trace(go.Scatter(
                x=[row["子市场"]],
                y=[top_y1],
                mode="markers+text",
                marker=dict(color="white", size=10, line=dict(color="#ff4d4f", width=2)),
                text=[f"{row['有干预样本数'] - row['有效干预数']:.0f}"],
                textposition="top center",
                textfont=dict(color="#ff4d4f", size=14, family="Arial Black"),
                showlegend=i == 0,
                hoverinfo='skip',
                name='无效干预样本数'
            ))
        for i, annotation in enumerate(fig['layout']['annotations']):
            annotation['font'] = dict(family="Microsoft YaHei", size=20, color="black")
        
        max_y = max(v_total) if v_total else 1000
        fig.update_yaxes(range=[0, max_y * 1.3], row=1, col=2)
        # --- 6. 整体布局美化 ---
        fig.update_layout(
            height=600,
            width=1100,
            template="plotly_white",
            barmode='group', # 簇状排列
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.2,
                xanchor="center",
                x=0.5
            ),
            margin=dict(t=40, b=100, l=50, r=50),
            font=dict(family="Microsoft YaHei", size=12)
        )

        # 优化坐标轴
        fig.update_yaxes(title_text="样本数量", row=1, col=2,range=[0, 900])
        fig.update_xaxes(tickangle=0, row=1, col=2)

        st.plotly_chart(fig, width='stretch')

    df_m = df_filtered_yes[df_filtered_yes["周数"] == select_week].groupby("子市场").agg({
        "单周干预偏差率": "mean",
        "环比干预偏差率": "mean",
        "主料mrpsku": "count"
    }).reset_index()
    df_m = df_m.rename(columns={"主料mrpsku": "SKU数量"})
    df_m = df_m.sort_values(by="SKU数量", ascending=True)
    col_bias_left, col_bias_right = st.columns([2,3])
    df_filtered_weeks_yes['单周干预偏差率']=df_filtered_weeks_yes['单周干预偏差率'].abs()
    df_filtered_weeks_yes['环比干预偏差率']=df_filtered_weeks_yes['环比干预偏差率'].abs()
    df_filtered_week_agg = df_filtered_weeks_yes.groupby("周数").agg({
        "单周干预偏差率": "mean",
        "环比干预偏差率": "mean"
    }).reset_index()
    with col_bias_left:
        st.markdown("#### 历史干预偏差趋势")
        fig_历史干预曲线=go.Figure()
        fig_历史干预曲线.add_trace(
            go.Scatter(
                x=df_filtered_week_agg["周数"],
                y=df_filtered_week_agg["单周干预偏差率"],
                mode='lines+markers+text',
                line=dict(color='#1f77b4', width=2),
                marker=dict(
                    color="#1f77b4", 
                    size=20, 
                    line=dict(width=5, color='white')
                ),
                text=df_filtered_week_agg["单周干预偏差率"].apply(lambda x: f"{x:.0%}"),
                textposition="top left",
                name="单周偏差",
                textfont=dict(size=14, color="black")
            )
        )
        fig_历史干预曲线.add_trace(
            go.Scatter(
                x=df_filtered_week_agg["周数"],
                y=df_filtered_week_agg["环比干预偏差率"],
                mode='lines+markers+text',
                line=dict(color='#ff7f0e', width=2),
                marker=dict(
                    color="#ff7f0e", 
                    size=20, 
                    line=dict(width=5, color='white')
                ),
                text=df_filtered_week_agg["环比干预偏差率"].apply(lambda x: f"{x:.0%}"),
                textposition="bottom left",
                name="环比偏差",
                yaxis="y2",
                textfont=dict(size=14, color="black")
            )
        )
        fig_历史干预曲线.add_hline(
            y=0.30,
            line_dash="dash",
            line_color="#1f77b4",  
            annotation_text="干预偏差: 30%",
            annotation_position="bottom right",
            opacity=0.7
            
        )
        fig_历史干预曲线.add_shape(
            type="line",
            xref="paper", yref="y2",
            x0=0, y0=0.05,
            x1=1, y1=0.05,
            line=dict(color="#ff7f0e", width=2, dash="dash"),
            opacity=0.7
        )
        max_y1 = df_filtered_week_agg["单周干预偏差率"].max().round(1)
        max_y2 = df_filtered_week_agg["环比干预偏差率"].max().round(1)
        fig_历史干预曲线.update_layout(
            # 左侧 Y 轴配置
            yaxis=dict(
                title="单周干预偏差率",
                range=[0, max_y1+0.2],  
                tickformat=".0%", 
                side="left"
            ),
            # 右侧 Y 轴配置
            yaxis2=dict(
                title="环比干预偏差率",
                range=[0, max_y2+0.05],  
                tickformat=".0%", 
                overlaying="y",   
                side="right"
            ),
            # 可选：优化整体布局，防止右侧 Y 轴标题被截断
            margin=dict(r=80),
            legend=dict(
                orientation="h",          # 水平排列
                yanchor="bottom",         # 垂直方向以底部为锚点
                y=1.02,                   # 放在绘图区顶部稍微偏上的位置
                xanchor="center",         # 水平方向以中心为锚点
                x=0.5,                    # 放在水平方向 50% 的位置（居中）
                bgcolor="rgba(255,255,255,0.8)" # 可选：设置半透明背景，防止遮挡标题或边框
            ),
            font=dict(family="Microsoft YaHei",size=12)
        )

        st.plotly_chart(fig_历史干预曲线,width="stretch",height=500)



    with col_bias_right:
        TARGET_MAX = 0.30
        st.markdown("#### 各市场干预偏差全貌 (目标区间: 30%)")

        # --- 2. 创建布局 ---
        fig = make_subplots(
            rows=1, cols=2, 
            shared_yaxes=True, 
            horizontal_spacing=0.03, 
            column_widths=[0.7, 0.3]
        )

        # --- 4. 左侧：单周干预偏差率 (条形图) ---
        # 颜色逻辑：出界的标红，在区间内的用青色
        colors_bar = [
            "#ff9d4f" if (x > TARGET_MAX) else "#5da9c4" 
            for x in df_m["单周干预偏差率"]
        ]

        fig.add_trace(
            go.Bar(
                y=df_m["子市场"],
                x=df_m["单周干预偏差率"],
                orientation='h',
                marker_color=colors_bar,
                text=df_m["单周干预偏差率"].apply(lambda x: f"{x:.1%}"),
                textposition='outside',
                cliponaxis=False,
                name="单周干预偏差",
                textfont=dict(size=14, color="black")
            ),
            row=1, col=1
        )

        for line_x in [TARGET_MAX]:
            fig.add_vline(
                x=line_x, 
                line_dash="dash", 
                line_color="red", 
                line_width=2,
                row=1, col=1,
                # 不透明度
                opacity=0.8
            )


        # --- 5. 右侧：环比干预偏差率 (折线图) ---
        colors_line = ["#5da9c4" if abs(x) <= 0.05 else "#f5222d" for x in df_m["环比干预偏差率"]]

        fig.add_trace(
            go.Scatter(
                y=df_m["子市场"],
                x=df_m["环比干预偏差率"],
                mode='lines+markers+text',
                line=dict(color='#bfbfbf', width=3),
                marker=dict(
                    color=colors_line, 
                    size=15, 
                    line=dict(width=5, color='white')
                ),
                text=df_m["环比干预偏差率"].apply(lambda x: f"{x:.0%}"),
                textposition="middle right",
                name="环比趋势",
                textfont=dict(size=14, color="black")
            ),
            row=1, col=2
        )
        for line_x in [0.05]:
            fig.add_vline(
                x=line_x, 
                line_dash="dash", 
                line_color="red", 
                line_width=2,
                row=1, col=2,
                # 不透明度
                opacity=0.8
            )

        max_val = max(df_m["单周干预偏差率"].max(), 0.5)
        min_val = min(df_m["单周干预偏差率"].min(), -0.5)

        fig.update_layout(
            template="simple_white",
            showlegend=False,
            height=len(df_m) * 45 + 120,
            margin=dict(l=10, r=60, t=20, b=40),
            xaxis=dict(
                title="干预偏差率",
                tickformat=".0%",
                range=[0, max_val+0.3], # 留出文本空间
                zeroline=True, zerolinecolor="#8c8c8c"
            ),
            xaxis2=dict(
                title="环比偏差率",
                tickformat=".0%",
                range=[0, 0.5], # 环比范围固定，方便观察斜率
                showgrid=False,
                zeroline=True,
                # 修改X轴区间范围
                # tickvals=[-1.1, -0.05, 0, 0.05, 1.1]
            ),
            font=dict(family="Microsoft YaHei"),
        )

        # 样式细节：隐藏右图Y轴刻度，统一字体
        fig.update_yaxes(showgrid=False, row=1, col=1)
        fig.update_yaxes(showticklabels=False, row=1, col=2)
        fig.update_yaxes(tickfont=dict(size=13), row=1, col=1)

        st.plotly_chart(fig, width='stretch',height=500)
    
    detail_cat,detail_sku_ganyu = st.columns([2,3])
    df_cat = df_filtered[df_filtered["周数"] == select_week].groupby("品类").agg({
        "单周干预偏差率": "mean",
        "环比干预偏差率": "mean",
        "主料mrpsku": "nunique" 
    }).reset_index()
    df_cat.columns = ["品类", "单周干预偏差率", "环比干预偏差率", "SKU个数"]
    sku_threshold = df_cat["SKU个数"].quantile(0.9)
    def get_label(row):
        if (abs(row["单周干预偏差率"]) > 0.3 or abs(row["环比干预偏差率"]) > 0.05) and row["SKU个数"] >= sku_threshold:
            return row["品类"]
        return ""

    df_cat["显示标签"] = df_cat.apply(get_label, axis=1)
    def get_color(row):
        if row["单周干预偏差率"] > 0.3: 
            return "干预过高"
        if row["单周干预偏差率"] < -0.3:
            return "干预过低"
        return "正常"
    
    df_cat["偏差情况"] = df_cat.apply(get_color, axis=1)
    df_cat['单周干预偏差率'] = round(df_cat['单周干预偏差率']*100, 1)
    df_cat['环比干预偏差率'] = round(df_cat['环比干预偏差率']*100, 1)
    with detail_cat:
        markdown_text,download_button = st.columns([3,1])
        markdown_text.markdown(
            f"""
            #### 干预情况下钻-<span style='color: #ff4b4b;'>{st.session_state.ganyu_filter_market}</span> 品类明细表
            """, 
            unsafe_allow_html=True
        )
        def color_deviation(val,target):
            color = '#cf1322' if abs(val) > target else '#389e0d'
            return f'color: {color}; font-weight: bold' if abs(val) > target else f'color: {color}'
        df_cat['周数']=select_week
        def to_excel(df):
            output = io.BytesIO()
            # 使用 xlsxwriter 作为引擎
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
            
            processed_data = output.getvalue()
            return processed_data
        excel_data = to_excel(df_cat)
        download_button.download_button(
            label="📥下载 品类明细表",
            data=excel_data,
            file_name=f'品类干预明细表_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        styled_df = df_cat[['周数','品类', '单周干预偏差率', '环比干预偏差率', 'SKU个数', '偏差情况']].sort_values("单周干预偏差率", ascending=False).style.map(
            color_deviation, subset=['单周干预偏差率'], target=30
        )
        styled_df = styled_df.map(
            color_deviation, subset=['环比干预偏差率'], target=5
        )


        st.dataframe(styled_df,
        column_config={
            "单周干预偏差率": st.column_config.NumberColumn(
                "单周干预偏差率(目标值:30%)",
                help="干预值与实际值的偏移程度",
                format="%.1f%%",
                width=80,
                alignment="center"
            ),
            "环比干预偏差率": st.column_config.NumberColumn(
                "环比干预偏差率(目标值:5%)",
                help="本周对比上周干预偏差的变化",
                format="%.1f%%",
                width=80,
                alignment="center"
            ),
            "SKU个数": st.column_config.NumberColumn("SKU个数", width=50, alignment="center"),
            "偏差情况": st.column_config.TextColumn("偏差情况", width=80),
            "品类": st.column_config.TextColumn("品类", width=50),
        },
         width="stretch", height=400, hide_index=True)
    
    with detail_sku_ganyu:
        markdown_text,select_shifouganyu,download_button = st.columns([3,1,1])
        markdown_text.markdown(
            f"""
            #### 干预情况下钻-<span style='color: #ff4b4b;'>{st.session_state.ganyu_filter_market}_{st.session_state.ganyu_filter_category}</span> MRPSKU明细表
            """, 
            unsafe_allow_html=True
        )
        
        inter_list = ["全部"] + sorted(df_filtered_yes["是否应该干预"].unique().tolist())
        with select_shifouganyu:
            selected_inter = st.selectbox("是否应该干预", inter_list,key="selectbox_inter")

        df_filtered_shifouganyu = df_filtered_yes.copy()
        if selected_inter != "全部":
            df_filtered_shifouganyu = df_filtered_yes[df_filtered_yes["是否应该干预"] == selected_inter]
        
        df_detail = df_filtered_shifouganyu[df_filtered_shifouganyu["周数"] == select_week][['周数','子市场', "channel_name",'品类', "主料mrpsku","状态", "当周实际值", "当周干预值", "单周干预偏差率", "环比干预偏差率", "是否应该干预"]].copy()
        df_detail['单周干预偏差率'] = round(df_detail['单周干预偏差率']*100, 1)
        df_detail['环比干预偏差率'] = round(df_detail['环比干预偏差率']*100, 1)
        with download_button:
            st.write(" ")
            st.write(" ")
            def to_excel(df):
                output = io.BytesIO()
                # 使用 xlsxwriter 作为引擎
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
                
                processed_data = output.getvalue()
                return processed_data
            excel_data = to_excel(df_detail)
            download_button.download_button(
                label="📥下载 MRPSKU明细表",
                data=excel_data,
                file_name=f'MRPSKU干预明细表_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )


        def color_deviation(val,target):
            color = '#cf1322' if abs(val) > target else '#389e0d'
            return f'color: {color}; font-weight: bold' if abs(val) > target else f'color: {color}'
        styled_df = df_detail[['周数','子市场', "channel_name",'品类', "主料mrpsku","状态", "当周实际值", "当周干预值", "单周干预偏差率", "环比干预偏差率", "是否应该干预"]].sort_values("单周干预偏差率", ascending=False).style.map(
            color_deviation, subset=['单周干预偏差率'],target=30
        )
        styled_df = styled_df.map(
            color_deviation, subset=['环比干预偏差率'], target=5
        )
        st.dataframe(
            styled_df,
            column_config={
                "周数": st.column_config.TextColumn("周数",width=50),
                "子市场": st.column_config.TextColumn("子市场",width=90),
                "channel_name": st.column_config.TextColumn("渠道",width=90),
                "品类": st.column_config.TextColumn("品类",width=80),
                "主料mrpsku": st.column_config.TextColumn("MRPSKU", width="medium"),
                "状态": st.column_config.TextColumn("状态", width=80),
                "当周实际值": st.column_config.NumberColumn("实际周销", format="%.0f", width=80, alignment="center"),
                "当周干预值": st.column_config.NumberColumn("干预周销", format="%.0f", width=80, alignment="center"),
                "单周干预偏差率": st.column_config.NumberColumn(
                    "单周干预偏差率(目标值:30%)",
                    help="干预值与实际值的偏移程度",
                    format="%.1f%%",
                    width=80,
                    alignment="center"
                ),
                "环比干预偏差率": st.column_config.NumberColumn(
                    "环比干预偏差率(目标值:5%)",
                    help="本周对比上周干预偏差的变化",
                    format="%.1f%%",
                    width=80,
                    alignment="center"
                ),
                "是否应该干预": st.column_config.TextColumn("是否应该干预", width=80)
            },
            hide_index=True,
            width="stretch",
            height=380
        )

@st.fragment
def delivery_stock_area(df_fahuo,df_历史海外周转,df_断货无在途, curr_filters):
    df = apply_filters(df_fahuo, curr_filters)
    df_历史海外周转 = apply_filters(df_历史海外周转, curr_filters)
    df_断货无在途 = apply_filters(df_断货无在途, curr_filters)
    if df is None or df.empty: return


    market_filter, category_filter, sku_filter,outStock_filter,week_filter = st.columns(5)
    with market_filter:
        market_list = ["全部市场"] + sorted(df["子市场"].unique().tolist())
        selected_market = st.selectbox("选择要查看的子市场", market_list,key="selectbox_market_fahuo")
        st.session_state.fahuo_filter_market=selected_market
    with category_filter:
        category_list = ["全部品类"] + sorted(df["品类"].unique().tolist())
        selected_category = st.selectbox("选择要查看的品类", category_list,key="selectbox_category_fahuo")
        st.session_state.fahuo_filter_category=selected_category
    with sku_filter:
        SKU_list = ["全部SKU"] + sorted(df["主料mrpsku"].unique().tolist())
        selected_sku = st.selectbox("选择要查看的SKU", SKU_list,key="selectbox_sku_fahuo")
        st.session_state.fahuo_filter_sku=selected_sku
    with outStock_filter:
        outStock_status_list = ["全部状态","断货","非断货"]
        select_stockStatus = st.selectbox("选择SKU的状态", outStock_status_list,key="selectbox_outStock_fahuo")
        st.session_state.fahuo_filter_outStockStatus=select_stockStatus
    with week_filter:
        if st.session_state.t0_date is not None:
            default_monday=st.session_state.t0_date.strftime("%Y-%m-%d")
        else:
            default_monday = (today - timedelta(days=today.weekday()) - timedelta(days=7)).strftime("%Y-%m-%d")
        select_week = st.date_input("选择要查看的周数", value=default_monday,key="date_input_week_fahuo")
        select_week = select_week.strftime("%Yw%V")
        st.session_state.fahuo_filter_week=select_week

    df_filtered = df.copy()
    if selected_market != "全部市场":
        df_filtered = df_filtered[df_filtered["子市场"] == selected_market]

    if selected_category != "全部品类":
        df_filtered = df_filtered[df_filtered["品类"] == selected_category]

    if selected_sku != "全部SKU":
        df_filtered = df_filtered[df_filtered["主料mrpsku"] == selected_sku]

    if select_stockStatus == "断货":
        df_filtered = df_filtered[df_filtered["状态"] == "断货"]
    elif select_stockStatus == "非断货":
        df_filtered = df_filtered[df_filtered["状态"] != "断货"]

    df_周转_filtered = df_历史海外周转.copy()
    if selected_market != "全部市场":
        df_周转_filtered = df_周转_filtered[df_周转_filtered["子市场"] == selected_market]

    if selected_category != "全部品类":
        df_周转_filtered = df_周转_filtered[df_周转_filtered["品类"] == selected_category]

    if selected_sku != "全部SKU":
        df_周转_filtered = df_周转_filtered[df_周转_filtered["主料mrpsku"] == selected_sku]

    if select_stockStatus == "断货":
        df_周转_filtered = df_周转_filtered[df_周转_filtered["状态"] == "断货"]
    elif select_stockStatus == "非断货":
        df_周转_filtered = df_周转_filtered[df_周转_filtered["状态"] != "断货"]

    df_断货无在途_filtered = df_断货无在途.copy()
    if selected_market != "全部市场":
        df_断货无在途_filtered = df_断货无在途_filtered[df_断货无在途_filtered["子市场"] == selected_market]
    if selected_category != "全部品类":
        df_断货无在途_filtered = df_断货无在途_filtered[df_断货无在途_filtered["品类"] == selected_category]
    if selected_sku != "全部SKU":
        df_断货无在途_filtered = df_断货无在途_filtered[df_断货无在途_filtered["主料mrpsku"] == selected_sku]



    df_filtered_oneweek = df_filtered[df_filtered["周数"]==select_week]
    df_filtered_weeks = df_filtered.copy()
    df_周转_filtered_oneweek = df_周转_filtered[df_周转_filtered["周数"]==select_week]
    df_周转_filtered_weeks = df_周转_filtered.copy()
    stage_cols = ["计划达成率", "配货达成率", "排单达成率", "出库达成率"]
    df_avg = df_filtered_oneweek.groupby("子市场").agg({
        "计划达成率": "mean",
        "配货达成率": "mean",
        "排单达成率": "mean",
        "出库达成率": "mean",
        "主料mrpsku": "count"
    }).reset_index()
    df_avg = df_avg.rename(columns={"主料mrpsku": "SKU数量"})
    df_avg = df_avg.sort_values(by="SKU数量", ascending=False)
    # 计算全市场平均值
    df_avg_allmarket = df_filtered_oneweek[stage_cols].mean().reset_index()
    df_avg_allmarket[0] = df_avg_allmarket[0].astype(float)
    
    col1, col2 = st.columns([1.5, 3])
    thresholds = {
        "计划达成率": 0.9, 
        "配货达成率": 0.9,
        "排单达成率": 1.0,
        "出库达成率": 1.0
    }
    def style_threshold_clean(col):
        if col.name in thresholds:
            target = thresholds[col.name]
            styles = []
            for v in col:
                if v >= target:
                    styles.append('color: #2E7D32;') 
                else:
                    styles.append('background-color: #FFF5F5; color: #C62828; font-weight: bold;')
            return styles
        return [''] * len(col)

    styled_df = (
        df_avg.style
        .apply(style_threshold_clean)
        .set_properties(**{
            'text-align': 'center',
        })
        .format("{:.1%}", subset=list(thresholds.keys()))
    )
    with col1:
        st.markdown("#### 📦 全市场发货过程指标全貌")
        c1,c2,c3,c4 = st.columns(4)
        # 如果指标值小于目标就为红色，反之为绿色
        c1.metric("计划达成率", f"{df_avg_allmarket.iloc[0,1]:.1%}",delta="目标: 90%",delta_color="green")
        c2.metric("配货达成率", f"{df_avg_allmarket.iloc[1,1]:.1%}",delta="目标: 90%",delta_color="green")
        c3.metric("排单达成率", f"{df_avg_allmarket.iloc[2,1]:.1%}",delta="目标: 100%",delta_color="green")
        c4.metric("出库达成率", f"{df_avg_allmarket.iloc[3,1]:.1%}",delta="目标: 100%",delta_color="green")
        st.dataframe(
            styled_df, 
            width='stretch', 
            column_config={
                "子市场": st.column_config.TextColumn("子市场", width="small"),
                **{
                    col: st.column_config.TextColumn(
                        label=f"{col}\n(目标:{thresholds[col]:.0%})", 
                        help=f"目标: {thresholds[col]:.1%}",
                        alignment="center"
                    ) for col in thresholds.keys()
                },
                "SKU数量": st.column_config.TextColumn("SKU数量", width="small",alignment="center"),
            },
            height=480,
            hide_index=True,
        )

    with col2:
        st.markdown("#### 📦 发货指标下钻分析-子市场")
        df_达成率 = df_filtered_weeks.groupby(['子市场', '周数'], as_index=False).agg(
            配货达成率=('配货达成率', 'mean')
        )
        df_SKU数量 = df_filtered_weeks.groupby(['子市场'], as_index=False).agg(
            SKU数量=('主料mrpsku', 'count')
        )
        df_达成率 = pd.merge(df_达成率, df_SKU数量, on=['子市场'], how='left')
        df_周转 = df_周转_filtered_weeks.copy()
        # 提前计算好带单价的金额，提升后续聚合速度
        df_周转['期初在库金额'] = df_周转['历史当周期初在库'] * df_周转['单价']
        df_周转['下周期初在库金额'] = df_周转['历史下周期初在库'] * df_周转['单价']
        df_周转['期初在途金额'] = df_周转['历史当周期初在途'] * df_周转['单价']
        df_周转['下周期初在途金额'] = df_周转['历史下周期初在途'] * df_周转['单价']
        df_周转['周销金额'] = df_周转['历史当周周销'] * df_周转['单价']

        exclude_mask = (
            (df_周转['状态'] == '清仓') | 
            (df_周转['状态'] == '新品FBM') | 
            ((df_周转['状态'] == '新品') & (df_周转['历史当周期初在库'] == 0))
        )
        df_周转['是否断货'] = np.where(exclude_mask, np.nan, df_周转['状态'] == '断货')
        # df_周转['是否断货'] = (df_周转['状态'] == '断货')
        df_周转_指标 = df_周转.groupby(['子市场', '周数'], as_index=False).agg(
            海外在库周转=('期初在库金额', lambda x: ((x.sum() + df_周转.loc[x.index, '下周期初在库金额'].sum()) / 2) / (df_周转.loc[x.index, '周销金额'].sum() / 7)),
            海外在途周转=('期初在途金额', lambda x: ((x.sum() + df_周转.loc[x.index, '下周期初在途金额'].sum()) / 2) / (df_周转.loc[x.index, '周销金额'].sum() / 7)),
            断货率=('是否断货', 'mean')
        )

        df_周转_指标['海外周转天数'] = df_周转_指标[['海外在库周转', '海外在途周转']].sum(axis=1)
        df_周转_指标[['海外在库周转', '海外在途周转', '海外周转天数']] = df_周转_指标[['海外在库周转', '海外在途周转', '海外周转天数']].round(1)

        result_df = pd.merge(df_周转_指标,df_达成率, on=['子市场', '周数'], how='outer')
        result_df=result_df.fillna(0)
        result_df['子市场总SKU'] = result_df.groupby('子市场')['SKU数量'].transform('sum')
        result_df = result_df.sort_values(
            by=['子市场总SKU', '周数', 'SKU数量'], 
            ascending=[False, True, False]
        )
        result_df = result_df.drop(columns=['子市场总SKU'])
        # 删除海外周转天数列值为0的行
        result_df = result_df[result_df['海外周转天数'] != 0]
        result_df['周数new'] = result_df["周数"].str[2:]
        fig_子市场发货 = make_subplots(
            rows=2, cols=1, 
            shared_xaxes=True,          
            vertical_spacing=0.04,
            row_heights=[0.35, 0.65], 
            specs=[[{"secondary_y": False}], [{"secondary_y": False}]] 
        )

        result_df['周数new'] = result_df["周数"].str[2:]
        # 不同子市场添加目标天数
        result_df['目标天数'] = np.where(
            result_df['子市场'] == 'DE',
            129,
            np.where(
                result_df['子市场'] == 'APM',
                135,
                np.where(
                    result_df['子市场'] == 'AP-CA',
                    135,
                    105
                )
            )
        )
        result_df['目标天数'] = result_df['目标天数'].round(0)
        result_df['子市场_new'] = (
            result_df['子市场'] + '<br>' +
            '(' + result_df['目标天数'].astype(str) + '天)'
        )
       
        fig_子市场发货.add_trace(
            go.Bar(
                x=[result_df['子市场_new'], result_df['周数new']], 
                y=result_df['海外在库周转'],
                name="海外在库周转",
                marker_color='#85C1E9',
                text=[f"{v:.0f}" for v in result_df['海外在库周转']],
                textposition='inside',
                textfont=dict(size=11, color="black"),
                cliponaxis=False,
                showlegend=True,
                hovertemplate=(
                    "子市场: %{x[0]}<br>"  # x[0] 对应 '子市场'
                    "周数: %{x[1]}<br>"    # x[1] 对应 '周数new'
                    "海外在库周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
                ),
                hoverlabel=dict(
                    font_size=10,
                    font_family="Microsoft YaHei",
                    font_color="black",
                    bgcolor="white"
                )
            ),
            row=2, col=1, secondary_y=False,
        )
        fig_子市场发货.add_trace(
            go.Bar(
                x=[result_df['子市场_new'], result_df['周数new']], 
                y=result_df['海外在途周转'],
                name="海外在途周转",
                marker_color='#f8c471',
                text=[f"{v:.0f}" for v in result_df['海外在途周转']],
                textposition='inside',
                textfont=dict(size=11, color="black"),
                cliponaxis=False,
                showlegend=True,
                hovertemplate=(
                    "子市场: %{x[0]}<br>"  # x[0] 对应 '子市场'
                    "周数: %{x[1]}<br>"    # x[1] 对应 '周数new'
                    "海外在途周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
                ),
                hoverlabel=dict(
                    font_size=10,
                    font_family="Microsoft YaHei",
                    font_color="black",
                    bgcolor="white"
                )
            ),
            row=2, col=1, secondary_y=False,
        )
        fig_子市场发货.add_trace(
            go.Scatter(
                x=[result_df['子市场_new'], result_df['周数new']],
                y=result_df['海外周转天数'],
                mode='text',
                text=[f"{v:.0f}" for v in result_df['海外周转天数']],
                textposition='top center',
                textfont=dict(
                    size=14,
                    color='#0461AB'
                ),
                showlegend=False,
                hoverinfo='skip'
            ),
            row=2, col=1, secondary_y=False
        )

        line_color = '#2E86C1'
        first_market = True # 用于控制图例只显示一次
        for market in result_df['子市场_new'].unique():
            df_sub = result_df[result_df['子市场_new'] == market]
            fig_子市场发货.add_trace(
                go.Scatter(
                    x=[df_sub['子市场_new'], df_sub['周数new']], 
                    y=df_sub['配货达成率'],
                    mode='lines+markers+text',
                    line=dict(color=line_color, width=2, shape='linear'),
                    marker=dict(color=line_color, size=6),
                    # 文本标签处理
                    text=[f"{v*100:.0f}%" if v==v else "" for v in df_sub['配货达成率']],
                    textposition="top center",
                    textfont=dict(size=13, color=line_color),
                    name="配货达成率",
                    legendgroup="配货达成率",      # 将所有子市场的线归为一组
                    showlegend=first_market,     # 只有第一次循环时显示图例
                    hoverinfo='skip',
                ),
                row=1, col=1
            )
            first_market = False


        top_y1 = (result_df['海外在库周转'].fillna(0) + result_df['海外在途周转'].fillna(0)).max() * 1.1
        # top_y1 = 430
        y_constant_list = [top_y1] * len(result_df)
        # 断货率如何大于0字体就是红色否则是绿色
        def get_trend_text_and_color(val):
            if val > 0:
                return "red"
            else:
                return "green"
        fig_子市场发货.add_trace(
            go.Scatter(
                x=[result_df['子市场_new'], result_df['周数new']],
                y=y_constant_list,
                mode="lines+markers+text",  
                line=dict(
                    color="#D3D3D3", 
                    width=1.5,             
                    dash="solid"           
                ),
                marker=dict(
                    # color="#66CC99", 
                    color = ["red" if v > 0 else "green" for v in result_df['断货率']],
                    size=6, 
                    line=dict(
                        # color="#66CC99",
                        color = ["red" if v > 0 else "green" for v in result_df['断货率']],
                        width=2
                    )
                ),
                # 正确提取单行对应的断货率文本
                text=[f"{v*100:.0f}%" if v==v else "" for v in result_df['断货率']],
                textposition="top center",
                textfont=dict(
                    # color="#009966", 
                    color = ["red" if v > 0 else "green" for v in result_df['断货率']],

                    size=14, 
                    family="Microsoft YaHei"
                ),
                showlegend=True,
                hoverinfo='skip',
                name='断货率'
            ),
            row=2, col=1, secondary_y=False
        )
        market_counts = result_df['子市场'].value_counts(sort=False)
        unique_markets = result_df['子市场'].unique()
        shapes = []
        current_position = -0.5
        for market in unique_markets[:-1]:
            current_position += market_counts[market]
            shapes.append(
                dict(
                    type="line",
                    xref="x2",
                    yref="paper",
                    x0=current_position,
                    x1=current_position,
                    y0=0,
                    y1=1,
                    line=dict(
                        color="#CCCCCC",  
                        width=2,
                        dash="dash"
                    )
                )
            )
        

        # ==================== 样式与坐标轴更新 ====================
        fig_子市场发货.update_layout(
            barmode='stack',
            height=650,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.05,
                xanchor="center",
                x=0.5,
                font=dict(size=14, color="black")
            ),
            plot_bgcolor='white',
            margin=dict(t=40, b=10, l=10, r=10),
            font=dict(family="Microsoft YaHei"),
            shapes=shapes
        )

        fig_子市场发货.update_xaxes(visible=False, row=1, col=1) 
        fig_子市场发货.update_yaxes(visible=False, range=[-0.05, 1.15], row=1, col=1, secondary_y=False)
        fig_子市场发货.update_yaxes(showgrid=False, zeroline=False, range=[0, (result_df['海外在库周转'].fillna(0) + result_df['海外在途周转'].fillna(0)).max() * 1.3], row=2, col=1, secondary_y=False)

        # 下方柱状图 X 轴
        fig_子市场发货.update_xaxes(
            tickangle=60,          
            showdividers=True,    
            dividercolor="#999999",
            row=2, col=1,
            tickfont=dict(size=12, color="black")
        )

        st.plotly_chart(fig_子市场发货, width='stretch')





    st.markdown(
        f"""
        #### 🔍发货指标下钻分析 - MRPSKU
        """, 
        unsafe_allow_html=True
    )

    total_skus = len(df_filtered_oneweek)
    problem_skus = len(df_filtered_oneweek[df_filtered_oneweek['实际出库量'] < df_filtered_oneweek['计划发货量']])
    # 用小组件显示概况
    m1, m2, m3,m4,m5 = st.columns(5)
    m1.metric("SKU 总数", total_skus)
    m2.metric("异常 SKU 数", problem_skus)
    with m3:
        st.write(" ")
        st.write("💡 *异常定义：实际出库量 < 计划发货量*")
    display_cols = [
        "子市场","主料mrpsku", "品类", "运输方式","计划发货量", "配货数量", "配货达成率","历史海外在库周转","历史海外在途周转","排单数量",  "实际出库量", "计划未达成原因"
    ]
    df_filtered_oneweek['缺口'] = df_filtered_oneweek['计划发货量'] - df_filtered_oneweek['实际出库量']
    df_filtered_oneweek['配货达成率'] = df_filtered_oneweek['配货达成率']*100
    df_filtered_sorted = df_filtered_oneweek.sort_values("缺口", ascending=False)
    exclude_mask = (
        (df_周转_filtered_oneweek['状态'] == '清仓') | 
        (df_周转_filtered_oneweek['状态'] == '新品FBM') | 
        ((df_周转_filtered_oneweek['状态'] == '新品') & (df_周转_filtered_oneweek['历史当周期初在库'] == 0))
    )
    df_周转_filtered_oneweek['用于计算断货的状态'] = np.where(exclude_mask, np.nan, (df_周转_filtered_oneweek['状态'] == '断货').astype(float))
    result_cat_df = df_周转_filtered_oneweek.groupby(
        ["子市场", "品类"],
        as_index=False
    ).agg(
        断货率=('用于计算断货的状态', 'mean')
    )
    df_周转_filtered_oneweek.drop(columns=['用于计算断货的状态'], inplace=True)
    
    df_filtered_sorted_merge = pd.merge(df_filtered_sorted, result_cat_df, on=["子市场","品类"], how="left")
    df_filtered_sorted_merge=df_filtered_sorted_merge[[
        "子市场","主料mrpsku", "品类","状态", "运输方式","计划发货量", "配货数量", "配货达成率","历史海外在库周转","历史海外在途周转","断货率","排单数量",  "实际出库量", "计划未达成原因"
    ]]

    with m4:
        reason_list = ['全部']+df_filtered_sorted_merge['计划未达成原因'].unique().tolist()
        selected_reason = st.selectbox("请选择计划未达成原因", reason_list)
        if selected_reason != '全部':
            df_filtered_sorted_merge = df_filtered_sorted_merge[df_filtered_sorted_merge['计划未达成原因'] == selected_reason]
    
    with m5:
        st.write(" ")
        st.write(" ")
        def to_excel(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
            processed_data = output.getvalue()
            return processed_data
        st.download_button(
            label="📥下载 发货指标下钻SKU明细",
            data=to_excel(df_filtered_sorted_merge),
            file_name=f"发货指标下钻SKU明细_{time.strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.ms-excel",
            key="download_excel"
        )
    
    def color_deviation1(val,target):
        color = '#389e0d' if abs(val) > target else '#cf1322'
        return f'color: {color}' if abs(val) > target else f'color: {color}; font-weight: bold'
    def color_deviation2(val,target):
        color = '#389e0d' if abs(val) <= target else '#cf1322'
        return f'color: {color}' if abs(val) <= target else f'color: {color}; font-weight: bold'
    styled_df = df_filtered_sorted_merge.sort_values("配货达成率", ascending=False).style.map(
        color_deviation1, subset=['配货达成率'],target=90
    )
    styled_df = styled_df.map(color_deviation2, subset=['断货率'],target=0)
    # 6. 展示表格
    st.dataframe(
        styled_df, 
        width='stretch', 
        height=500,
        column_config={
            "子市场": st.column_config.TextColumn("子市场", width=0.5),
            "品类": st.column_config.TextColumn("品类", width=0.5),
            "主料mrpsku": st.column_config.TextColumn("主料mrpsku", width=5),
            "状态": st.column_config.TextColumn("状态", width=0.5),
            "运输方式": st.column_config.TextColumn("运输方式", width=0.1),
            "计划发货量": st.column_config.NumberColumn("计划发货量", format="%d", alignment="center",width=1),
            "配货数量": st.column_config.NumberColumn("配货数量", format="%d", alignment="center",width=1),
            "配货达成率": st.column_config.NumberColumn("配货达成率", width=0.1,format="%.1f%%",alignment="center"),
            "历史海外在库周转": st.column_config.NumberColumn("海外在库周转", format="%d", alignment="center",width=1),
            "历史海外在途周转": st.column_config.NumberColumn("海外在途周转", format="%d", alignment="center",width=1),
            "断货率": st.column_config.NumberColumn("断货率", width=0.1,format="%.1f%%",alignment="center"),
            "排单数量": st.column_config.NumberColumn("排单数量", format="%d", alignment="center",width=1),
            "实际出库量": st.column_config.NumberColumn("实际出库量", format="%d", alignment="center",width=1),
            "计划未达成原因": st.column_config.TextColumn("计划未达成原因", width=350),
        },
        hide_index=True
    )

    #================= 断货无在途区域 ============================
    markdown_text,download_button=st.columns([3,1])
    with markdown_text:
        st.markdown("### 断货无在途SKU明细")
    with download_button:
        st.write(" ")
        st.write(" ")
        def to_excel(df):
            output = io.BytesIO()
            # 使用 xlsxwriter 作为引擎
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # 如果需要对 Excel 进行样式处理，可以在这里操作 writer
            
            processed_data = output.getvalue()
            return processed_data
        excel_data = to_excel(df_断货无在途_filtered)
        st.download_button(
            label="📥下载 断货无在途 数据",
            data=excel_data,
            file_name=f'断货无在途数据_{time.strftime("%Y-%m-%d", time.localtime())}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    st.dataframe(
        df_断货无在途_filtered[['子市场','主料mrpsku','层级','品类','二级品类','状态','二级状态','货源地','规格','可用库存','待上架库存','IQC']],
        width='stretch', 
        height=500,
        column_config={
            "子市场": st.column_config.TextColumn("子市场", width=0.5),
            "主料mrpsku": st.column_config.TextColumn("主料mrpsku", width=5),
            "层级": st.column_config.TextColumn("层级", width=0.5),
            "品类": st.column_config.TextColumn("品类", width=0.5),
            "二级品类": st.column_config.TextColumn("二级品类", width=0.5),
            "状态": st.column_config.TextColumn("状态", width=0.5),
            "二级状态": st.column_config.TextColumn("二级状态", width=0.5),
            "货源地": st.column_config.TextColumn("货源地", width=0.5),
            "规格": st.column_config.TextColumn("规格", width=0.5),
            "可用库存": st.column_config.NumberColumn("可用库存", format="%d", alignment="center",width=1),
            "待上架库存": st.column_config.NumberColumn("待上架库存", format="%d", alignment="center",width=1),
            "IQC": st.column_config.NumberColumn("IQC", format="%d", alignment="center",width=1),
        }
    )

    
@st.fragment
def actual_turnover_area(df_country_turnover,df_历史海外周转,filters):
    # df_country_turnover=apply_filters(df_country_turnover,filters)
    df_历史海外周转=apply_filters(df_历史海外周转,filters)
    if df_country_turnover is None or df_country_turnover.empty:
        st.warning("无数据")
        return

    df_ct = df_country_turnover.copy()
    df_ct['stock_val'] = df_ct['当周期初在库'] * df_ct['单价']
    df_ct['next_stock_val'] = df_ct['下周期初在库'] * df_ct['单价']
    df_ct['sales_val'] = df_ct['当周周销'] * df_ct['单价']

    # 一次性聚合，避免 Python 循环
    res_ct = df_ct.groupby('周数').agg({
        'stock_val': 'sum',
        'next_stock_val': 'sum',
        'sales_val': 'sum'
    })
    res_ct['国内在库周转'] = ((res_ct['stock_val'] + res_ct['next_stock_val']) / 2) / (res_ct['sales_val'] / 7)

    # --- 优化后的海外周转计算 ---
    df_hw = df_历史海外周转.copy()
    df_hw['stock_val'] = df_hw['历史当周期初在库'] * df_hw['单价']
    df_hw['next_stock_val'] = df_hw['历史下周期初在库'] * df_hw['单价']
    df_hw['transit_val'] = df_hw['历史当周期初在途'] * df_hw['单价']
    df_hw['next_transit_val'] = df_hw['历史下周期初在途'] * df_hw['单价']
    df_hw['sales_val'] = df_hw['历史当周周销'] * df_hw['单价']
    df_hw['is_duanhua'] = (df_hw['状态'] == '断货').astype(int)
    res_hw = df_hw.groupby('周数').agg({
        'stock_val': 'sum', 'next_stock_val': 'sum',
        'transit_val': 'sum', 'next_transit_val': 'sum',
        'sales_val': 'sum', 'is_duanhua': 'mean'
    })

    res_hw['海外在库周转'] = (((res_hw['stock_val'] + res_hw['next_stock_val']) / 2) / (res_hw['sales_val'] / 7)).round(1)
    res_hw['海外在途周转'] = (((res_hw['transit_val'] + res_hw['next_transit_val']) / 2) / (res_hw['sales_val'] / 7)).round(1)
    res_hw['断货率'] = res_hw['is_duanhua'] * 100
    # 合并结果
    result_df = pd.concat([res_ct['国内在库周转'], res_hw[['海外在库周转', '海外在途周转', '断货率']]], axis=1).reset_index()
    result_df['海外总周转'] = (((result_df['海外在库周转'] + result_df['海外在途周转']).round(1)))
    fig_历史周转 = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True, 
        vertical_spacing=0.1, # 上下两个图的间距
        row_heights=[0.3, 0.7]
    )
    fig_历史周转.add_trace(go.Bar(
            x=result_df['周数'],
            y=result_df['海外在库周转'],
            name='海外在库周转',
            text=result_df['海外在库周转'].apply(lambda x: f"{x:.1f}"),
            textposition="outside",
            marker=dict(
                color='#85C1E9'
            ),
            textfont=dict(size=18, color="black"),
            hovertemplate=(
                    "周数: %{x}<br>"    # x[1] 对应 '周数new'
                    "海外在库周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
            ),
            hoverlabel=dict(
                font_size=12,
                font_family="Microsoft YaHei",
                font_color="black",
                bgcolor="white"
            )
        ),
        row=2,
        col=1
    )
    fig_历史周转.add_trace(go.Bar(
            x=result_df['周数'],
            y=result_df['海外在途周转'],
            name='海外在途周转',
            text=result_df['海外在途周转'].apply(lambda x: f"{x:.1f}"),
            textposition="outside",
            marker=dict(
                color='#f8c471'
            ),
            textfont=dict(size=18, color="black"),
            hovertemplate=(
                    "周数: %{x}<br>"    # x[1] 对应 '周数new'
                    "海外在途周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
            ),
            hoverlabel=dict(
                font_size=12,
                font_family="Microsoft YaHei",
                font_color="black",
                bgcolor="white"
            )
        ),
        row=2,
        col=1
    )
    fig_历史周转.add_trace(go.Bar(
            x=result_df['周数'],
            y=result_df['国内在库周转'],
            name='国内在库周转',
            text=result_df['国内在库周转'].apply(lambda x: f"{x:.1f}"),
            textposition="outside",
            marker=dict(
                color='#bfc9ca'
            ),
            textfont=dict(size=18, color="black"),
            hovertemplate=(
                    "周数: %{x}<br>"    # x[1] 对应 '周数new'
                    "国内在库周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
            ),
            hoverlabel=dict(
                font_size=12,
                font_family="Microsoft YaHei",
                font_color="black",
                bgcolor="white"
            )
        ),
        row=2,
        col=1
    )
    fig_历史周转.add_trace(go.Scatter(
            x=result_df['周数'],
            y=result_df['海外总周转'],
            mode='lines+markers+text',
            name='海外总周转',
            text=result_df['海外总周转'].apply(lambda x: f"{x:.1f}"),
            textposition="top center",
            marker=dict(
                color='#2E8421',
                size=12
            ),
            textfont=dict(
                size=15,
                color="black",
            )
        ),
        row=2,
        col=1
    )
    fig_历史周转.add_hline(
            y=105,
            line_dash="dash",  # 设置为虚线
            line_color="#CC0033",  # 保持颜色一致
            annotation_text="海外周转目标: 105天",
            annotation_position="bottom right",  # 文字显示位置
            opacity=0.7  # 设置透明度，避免抢了主数据的视觉焦点
    )


    def get_trend_text_and_color(val):
        if val > 0:
            return f"↗{val:.1f}%", "red"
        else:
            return f"↘{abs(val):.1f}%", "green"

    text_labels = result_df['断货率'].apply(lambda x: get_trend_text_and_color(x)[0])
    marker_colors = result_df['断货率'].apply(lambda x: get_trend_text_and_color(x)[1])

    fig_历史周转.add_trace(
        go.Scatter(
            x=result_df['周数'],
            y=result_df['断货率'],
            mode='markers+text',
            text=text_labels,
            textposition="top center",
            textfont=dict(color=marker_colors, size=16),
            marker=dict(
                color=marker_colors,
                size=15,
                line=dict(width=1, color='white')
            ),
            # 关键：用error_y实现棒棒糖的垂直杆
            error_y=dict(
                type='data',
                symmetric=False,
                array=[0] * len(result_df), # 向上延伸量为0
                arrayminus=result_df['断货率'], # 向下延伸到0点
                width=0, # 不显示横向的小横杠
                thickness=1.5,
                color='rgba(100, 100, 100, 0.5)' # 杆子的颜色（浅灰色半透明）
            ),
            showlegend=True,
            name='断货率',
            hovertemplate='周数: %{x}<br>断货率: %{y}%<extra></extra>'
        ),
        row=1, col=1
    )

    # --- 全局布局调整 ---
    fig_历史周转.update_layout(
        barmode='group',
        height=600,
        plot_bgcolor='white', # 设置背景为白色更接近原图
        margin=dict(t=80, b=50, l=50, r=50),
        font=dict(family="Microsoft YaHei"),
        legend=dict(
            orientation="h",      
            yanchor="bottom",    
            y=1.02,               
            xanchor="center",     
            x=0.5,                
            title_text="",
            font=dict(size=16, color="black")
        )
    )
    max_y = max(result_df['海外总周转']) + 30
    # --- 坐标轴设置 ---
    fig_历史周转.update_yaxes(
        title_text="周转天数",
        showgrid=True,
        gridcolor='lightgray',
        row=2, col=1,
        tickfont=dict(size=16, color="black"),
        range = [0 , max_y]
    )

    fig_历史周转.update_yaxes(
        range=[0, 3], # 稍微给上方文本留点空间
        showticklabels=False, # 隐藏数值标签
        showgrid=False,       # 隐藏网格线
        zeroline=True,        # 显示 0 基准线
        zerolinecolor='gray',
        zerolinewidth=1,
        row=1, col=1
    )

    # 调整下方 X 轴
    fig_历史周转.update_xaxes(
        type='category', # 确保周数按类别等距显示
        row=2, col=1,
        tickfont=dict(size=16, color="black")
    )

    st.plotly_chart(fig_历史周转, width='stretch')
        

# =========================================================
# 6、主体渲染
# =========================================================
if st.session_state.df_fahuo is not None:
    curr_filters = st.session_state.committed_filters
    # 历史实际周转区域
    st.header("📈 历史实际周转", anchor="0")
    actual_turnover_area(st.session_state.df_country_turnover,st.session_state.df_历史海外周转, curr_filters)
    st.divider()
    
    # 库销比区域
    st.header("📉 海外库存周转", anchor="1")
    inventorySales_rate_area(st.session_state.df_stock_turnover,st.session_state.df_历史海外周转, curr_filters)
    st.divider()
    # 发货指标区域
    st.markdown("# 供")
    st.header("🚚 发货过程指标", anchor="2")
    delivery_stock_area(st.session_state.df_fahuo,st.session_state.df_历史海外周转,st.session_state.df_断货无在途, curr_filters)
    st.divider()
    # 预测指标区域
    st.markdown("# 销")
    st.header("📈 预测指标", anchor="3")
    predictSales_rate_area(st.session_state.df_yuce, curr_filters)
    st.divider()
    # 干预指标区域
    st.header("🛠️ 干预指标", anchor="4")
    ganyuSales_rate_area(st.session_state.df_ganyu, curr_filters)
else:
    st.info("👋 请先在左侧侧边栏上传数据文件。")


