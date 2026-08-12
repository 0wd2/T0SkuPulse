import streamlit as st
import plotly.graph_objects as go
import polars as pl
import pandas as pd
from io import BytesIO
from datetime import date, timedelta
from openpyxl import load_workbook
from plotly.subplots import make_subplots
import plotly.express as px
import numpy as np
import warnings
import fastexcel

warnings.filterwarnings("ignore") 

import io
import time
from xlsxwriter import Workbook

st.set_page_config(page_title="全层级指标监控", layout="wide", page_icon="📊")
st.markdown("<h2 style='text-align: center; color: #1E3A8A; margin:0;'>📊 全层级SKU指标监控</h2>",
                unsafe_allow_html=True)
if "df_海外周转" not in st.session_state: st.session_state.df_海外周转 = None
if "df_国内在库周转" not in st.session_state: st.session_state.df_国内在库周转 = None
if "df_断货率" not in st.session_state: st.session_state.df_断货率 = None

@st.cache_data(show_spinner="正在解析并格式化数据，请稍候...", ttl=3600)
def process_uploaded_files(uploaded_files):
    data_pool = {
        "df_海外周转": None, "df_国内在库周转": None,
        "df_断货率": None,
    }

    for uploaded_file in uploaded_files:
        file_name = uploaded_file.name
        file_bytes = uploaded_file.getvalue()
        file_type = file_name.split('.')[-1].lower()

        try:
            if file_type == 'xlsx':
                # 使用 BytesIO 读取
                with io.BytesIO(file_bytes) as f:
                    # 获取所有 Sheet 名称
                    workbook = load_workbook(f, read_only=True)
                    sn = workbook.sheetnames

                    # 定义内部格式化工具
                    def format_df(df):
                        if df is not None and not df.empty:
                            if '日期' in df.columns:
                                df['日期'] = pd.to_datetime(df['日期'])
                            return df
                        return None

                    # 批量读取各 Sheet
                    if '海外周转汇总' in sn:
                        data_pool["df_海外周转"] = format_df(pl.read_excel(f, sheet_name='海外周转汇总').to_pandas())
                    if '国内在库周转' in sn:
                        data_pool["df_国内在库周转"] = format_df(pl.read_excel(f, sheet_name='国内在库周转').to_pandas())
                    if '断货率' in sn:
                        data_pool["df_断货率"] = format_df(pl.read_excel(f, sheet_name='断货率').to_pandas())

            elif file_type == 'parquet':
                with io.BytesIO(file_bytes) as f:
                    data_pool["stock_turnover"] = pd.read_parquet(f)

        except Exception as e:
            st.error(f"解析文件 {file_name} 时出错: {e}")

    return data_pool


# 3、侧边栏配置
with st.sidebar:
    st.header("⚙️ 配置参数")
    today = date.today()
    default_monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
    t0_date_val = st.date_input("🗓️ 当前周周一", value=default_monday)
    st.session_state.t0_date = t0_date_val

    up_folder_btn = st.file_uploader("请选择上传文件：", type=['xlsx', 'parquet'], accept_multiple_files=True)

    if up_folder_btn:
        # 调用缓存函数处理文件
        with st.spinner("正在快速加载数据..."):
            processed_data = process_uploaded_files(up_folder_btn)

            # 将处理后的数据更新到 session_state（仅在有值时更新，避免覆盖已有的其他数据）
            for key, df in processed_data.items():
                if df is not None:
                    st.session_state[key] = df

            st.success("✅ 数据加载完成 (已缓存)")

    st.divider()


def 整体指标分层级(result_df,targetDays1,targetDays2,outStockRate):
    result_df = result_df.sort_values(by='周数')
    fig_历史周转 = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.1,
        row_heights=[0.3, 0.7]
    )
    fig_历史周转.add_trace(go.Bar(
        x=result_df['周数'],
        y=result_df['海外在库周转'],
        name='海外在库周转',
        text=result_df['海外在库周转'].apply(lambda x: f"{x:.1f}"),
        textposition="outside",
        marker=dict(
            color='#85C1E9'
        ),
        textfont=dict(size=18, color="black"),
        hovertemplate=(
            "周数: %{x}<br>"
            "海外在库周转: %{y:.0f}<extra></extra>"
        ),
        hoverlabel=dict(
            font_size=12,
            font_family="Microsoft YaHei",
            font_color="black",
            bgcolor="white"
        )
    ),
        row=2,
        col=1
    )
    fig_历史周转.add_trace(go.Bar(
        x=result_df['周数'],
        y=result_df['海外在途周转'],
        name='海外在途周转',
        text=result_df['海外在途周转'].apply(lambda x: f"{x:.1f}"),
        textposition="outside",
        marker=dict(
            color='#f8c471'
        ),
        textfont=dict(size=18, color="black"),
        hovertemplate=(
            "周数: %{x}<br>"
            "海外在途周转: %{y:.0f}<extra></extra>"
        ),
        hoverlabel=dict(
            font_size=12,
            font_family="Microsoft YaHei",
            font_color="black",
            bgcolor="white"
        )
    ),
        row=2,
        col=1
    )
    fig_历史周转.add_trace(go.Bar(
        x=result_df['周数'],
        y=result_df['国内在库周转天数'],
        name='国内在库周转',
        text=result_df['国内在库周转天数'].apply(lambda x: f"{x:.1f}"),
        textposition="outside",
        marker=dict(
            color='#bfc9ca'
        ),
        textfont=dict(size=19, color="black"),
        hovertemplate=(
            "周数: %{x}<br>"
            "国内在库周转: %{y:.0f}<extra></extra>"
        ),
        hoverlabel=dict(
            font_size=12,
            font_family="Microsoft YaHei",
            font_color="black",
            bgcolor="white"
        )
    ),
        row=2,
        col=1
    )
    fig_历史周转.add_trace(go.Scatter(
        x=result_df['周数'],
        y=result_df['海外周转天数'],
        mode='lines+markers+text',
        name='海外周转天数',
        text=result_df['海外周转天数'].apply(lambda x: f"{x:.1f}"),
        textposition="top center",
        marker=dict(
            color='#2E8421',
            size=12
        ),
        textfont=dict(
            size=15,
            color="black",
        )
    ),
        row=2,
        col=1
    )
    fig_历史周转.add_hline(
        y=targetDays1,
        line_dash="dash",
        line_color="#CC0033", 
        annotation_text=f"目标: {targetDays1}天",
        annotation_position="bottom right",  
        opacity=0.7
    )
    if targetDays2 != 0:
        fig_历史周转.add_hline(
            y=targetDays2,
            line_dash="dash",
            line_color="#CC0033", 
            annotation_text=f"目标: {targetDays2}天",
            annotation_position="bottom right",  
            opacity=0.7
        )

    def get_trend_text_and_color(val):
        if val > outStockRate:
            return f"{val:.2f}%", "red"
        else:
            return f"{abs(val):.2f}%", "green"

    text_labels = result_df['断货率'].apply(lambda x: get_trend_text_and_color(x)[0])
    marker_colors = result_df['断货率'].apply(lambda x: get_trend_text_and_color(x)[1])

    fig_历史周转.add_trace(
        go.Scatter(
            x=result_df['周数'],
            y=result_df['断货率'],
            mode='markers+text',
            text=text_labels,
            textposition="top center",
            textfont=dict(color=marker_colors, size=16),
            marker=dict(
                color=marker_colors,
                size=15,
                line=dict(width=1, color='white')
            ),
            error_y=dict(
                type='data',
                symmetric=False,
                array=[0] * len(result_df),  
                arrayminus=result_df['断货率'],
                width=0,
                thickness=1.5,
                color='rgba(100, 100, 100, 0.5)'
            ),
            showlegend=True,
            name='断货率',
            hovertemplate='周数: %{x}<br>断货率: %{y}%<extra></extra>'
        ),
        row=1, col=1
    )

    fig_历史周转.update_layout(
        barmode='group',
        height=600,
        plot_bgcolor='white',
        margin=dict(t=80, b=50, l=50, r=50),
        font=dict(family="Microsoft YaHei"),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5,
            title_text="",
            font=dict(size=16, color="black")
        )
    )
    max_y = max(result_df['海外周转天数']) + 30
    fig_历史周转.update_yaxes(
        title_text="周转天数",
        showgrid=True,
        gridcolor='lightgray',
        row=2, col=1,
        tickfont=dict(size=16, color="black"),
        range=[0, max_y]
    )
    max_out_of_stock = result_df['断货率'].max()
    upper_limit = max_out_of_stock * 1.5 if max_out_of_stock > 0 else 1 
    fig_历史周转.update_yaxes(
        range=[0, upper_limit],
        showticklabels=False,
        showgrid=False,
        zeroline=True,
        zerolinecolor='gray',
        zerolinewidth=1,
        row=1, col=1
    )

    fig_历史周转.update_xaxes(
        type='category',
        row=2, col=1,
        tickfont=dict(size=16, color="black")
    )
    return fig_历史周转




@st.fragment
def actual_turnover_area(df_海外周转, df_国内在库周转, df_断货率):
    if df_海外周转 is None or df_海外周转.empty:
        st.warning("无数据")    
        return

    df_hw = df_海外周转.groupby(['周数','层级']).agg(
         期初库存金额 = ("期初库存金额", "sum"),
         期末库存金额 = ("期末库存金额", "sum"),
         期初在途金额 = ("期初在途金额", "sum"),
         期末在途金额 = ("期末在途金额", "sum"),
         周销售出库金额 = ("周销售出库金额", "sum"),
    ).reset_index()
    df_hw['海外在库周转'] = ((df_hw['期末库存金额'] + df_hw['期初库存金额'])/2 / (df_hw['周销售出库金额']/7+0.001)).round(2)
    df_hw['海外在途周转'] = ((df_hw['期末在途金额'] + df_hw['期初在途金额'])/2 / (df_hw['周销售出库金额']/7+0.001)).round(2)
    # 合并结果
    df_nostock = df_断货率.groupby(['周数','层级']).agg(
        SKU总数 = ("sku总数", "sum"),
        断货SKU数量 = ("断货sku数量", "sum"),
    ).reset_index()
    df_nostock['断货率'] = df_nostock['断货SKU数量']/df_nostock['SKU总数'].round(4)
    df_nostock['断货率'] = df_nostock['断货率']*100
    result_df = pd.merge(df_国内在库周转[['周数','层级','国内在库周转天数']], df_hw, on=['周数','层级'], how='left')
    result_df = pd.merge(result_df, df_nostock, on=['周数','层级'], how='left')
    result_df['海外周转天数'] = (((result_df['海外在库周转'] + result_df['海外在途周转']).round(1)))
    result_df = result_df.sort_values(by=['周数','层级'])
    result_df = result_df.reset_index(drop=True)
    fig_历史周转_TOP0 = 整体指标分层级(result_df[result_df['层级']=='TOP0'],90,0,1)
    fig_历史周转_TOP1 = 整体指标分层级(result_df[result_df['层级']=='TOP1'],90,105,2)
    fig_历史周转_TOP10 = 整体指标分层级(result_df[result_df['层级']=='TOP10'],90,105,4)
    fig_历史周转_TOP20 = 整体指标分层级(result_df[result_df['层级']=='TOP20'],90,105,6)
    fig_历史周转_普通 = 整体指标分层级(result_df[result_df['层级']=='普通'],90,105,7.4)
    return fig_历史周转_TOP0, fig_历史周转_TOP1, fig_历史周转_TOP10, fig_历史周转_TOP20, fig_历史周转_普通

def 子市场指标分层级(result_df,targetDays1,targetDays2,outStockRate):
    mc_level = result_df['层级'].unique()[0]
    df_sorted_temp = result_df.sort_values(['子市场', '周数'])
    latest_sku_map = df_sorted_temp.groupby('子市场')['SKU数量'].last()
    result_df['最新SKU数量'] = result_df['子市场'].map(latest_sku_map)
    market_order = latest_sku_map.sort_values(ascending=False).index.tolist()
    result_df['SKU数量占比'] = (result_df['子市场'].map(latest_sku_map)/ latest_sku_map.sum())*100
    result_df['SKU数量占比'] = result_df['SKU数量占比'].round(1)
    result_df['子市场'] = pd.Categorical(result_df['子市场'], categories=market_order, ordered=True)
    result_df = result_df.sort_values(by=['子市场', '周数'])
    result_df = result_df.reset_index(drop=True)
    fig_子市场指标 = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.04,
        row_heights=[0, 1.0],
        specs=[[{"secondary_y": False}], [{"secondary_y": False}]]
    )
    result_df['周数new'] = result_df["周数"].str[2:]
    # # 不同子市场添加目标天数
    # if mc_level == 'TOP0':
    #     result_df['目标天数'] = np.where(
    #         result_df['子市场'] == 'DE',
    #         f"{targetDays1+24}",
    #         np.where(
    #             result_df['子市场'] == 'APM',
    #             f"{targetDays1+30}",
    #             np.where(
    #                 result_df['子市场'] == 'AP-CA',
    #                 f"{targetDays1+30}",
    #                 f"{targetDays1}"
    #             )
    #         )
    #     )
    # else:
    #     result_df["目标天数"] = np.where(
    #         result_df["子市场"] == "DE",
    #         f"{targetDays1 + 24}-{targetDays2 + 24}",
    #         np.where(
    #             result_df["子市场"] == "APM",
    #             f"{targetDays1 + 30}-{targetDays2 + 30}",
    #             np.where(
    #                 result_df["子市场"] == "AP-CA",
    #                 f"{targetDays1 + 30}-{targetDays2 + 30}",
    #                 f"{targetDays1}-{targetDays2}",
    #             ),
    #         ),
    #     )
    # # result_df['目标天数'] = result_df['目标天数'].round(0)
    result_df["子市场_new"] = (
        result_df["子市场"].astype(str)
        + "<br>"
        + "( "
        + result_df["最新SKU数量"].astype(str)
        + " 个)"
        + "<br>"
        + "( "
        + result_df["SKU数量占比"].astype(str)
        + " %)"
        + "<br>"
    )

    fig_子市场指标.add_trace(
        go.Bar(
            x=[result_df['子市场_new'], result_df['周数new']],
            y=result_df['海外在库周转'],
            name="海外在库周转",
            marker_color='#85C1E9',
            text=[f"{v:.0f}" for v in result_df['海外在库周转']],
            textposition='inside',
            textfont=dict(size=11, color="black"),
            cliponaxis=False,
            showlegend=True,
            hovertemplate=(
                "子市场: %{x[0]}<br>"  # x[0] 对应 '子市场'
                "周数: %{x[1]}<br>"  # x[1] 对应 '周数new'
                "海外在库周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
            ),
            hoverlabel=dict(
                font_size=10,
                font_family="Microsoft YaHei",
                font_color="black",
                bgcolor="white"
            )
        ),
        row=2, col=1, secondary_y=False,
    )
    fig_子市场指标.add_trace(
        go.Bar(
            x=[result_df['子市场_new'], result_df['周数new']],
            y=result_df['海外在途周转'],
            name="海外在途周转",
            marker_color='#f8c471',
            text=[f"{v:.0f}" for v in result_df['海外在途周转']],
            textposition='inside',
            textfont=dict(size=11, color="black"),
            cliponaxis=False,
            showlegend=True,
            hovertemplate=(
                "子市场: %{x[0]}<br>"  # x[0] 对应 '子市场'
                "周数: %{x[1]}<br>"  # x[1] 对应 '周数new'
                "海外在途周转: %{y:.0f}<extra></extra>"  # extra 用于去掉默认 trace 名称
            ),
            hoverlabel=dict(
                font_size=10,
                font_family="Microsoft YaHei",
                font_color="black",
                bgcolor="white"
            )
        ),
        row=2, col=1, secondary_y=False,
    )
    fig_子市场指标.add_trace(
        go.Scatter(
            x=[result_df['子市场_new'], result_df['周数new']],
            y=result_df['海外周转天数'],
            mode='text',
            text=[f"{v:.0f}" for v in result_df['海外周转天数']],
            textposition='top center',
            textfont=dict(
                size=14,
                color='#0461AB'
            ),
            showlegend=False,
            hoverinfo='skip'
        ),
        row=2, col=1, secondary_y=False
    )
    top_y1 = (result_df['海外在库周转'].fillna(0) + result_df['海外在途周转'].fillna(0)).max() * 1.1
    # top_y1 = 430
    y_constant_list = [top_y1] * len(result_df)

    # 断货率如何大于0字体就是红色否则是绿色
    def get_trend_text_and_color(val):
        if val > outStockRate:
            return "red"
        else:
            return "green"
    if mc_level == '普通':
        fig_子市场指标.add_trace(
            go.Scatter(
                x=[result_df['子市场_new'], result_df['周数new']],
                y=y_constant_list,
                mode="lines+markers+text",
                line=dict(
                    color="#D3D3D3",
                    width=1.5,
                    dash="solid"
                ),
                marker=dict(
                    # color="#66CC99",
                    color=["red" if v > outStockRate else "green" for v in result_df['断货率']],
                    size=6,
                    line=dict(
                        # color="#66CC99",
                        color=["red" if v > outStockRate else "green" for v in result_df['断货率']],
                        width=2
                    )
                ),
                # 正确提取单行对应的断货率文本
                text=[f"{v * 100:.0f}%" if v == v else "" for v in result_df['断货率']],
                textposition="top center",
                textfont=dict(
                    # color="#009966",
                    color=["red" if v > outStockRate else "green" for v in result_df['断货率']],

                    size=12,
                    family="Microsoft YaHei"
                ),
                showlegend=True,
                hoverinfo='skip',
                name='断货率'
            ),
            row=2, col=1, secondary_y=False
        )
    else:
        fig_子市场指标.add_trace(
            go.Scatter(
                x=[result_df['子市场_new'], result_df['周数new']],
                y=y_constant_list,
                mode="lines+markers+text",
                line=dict(
                    color="#D3D3D3",
                    width=1.5,
                    dash="solid"
                ),
                marker=dict(
                    # color="#66CC99",
                    color=["red" if v > outStockRate else "green" for v in result_df['断货率']],
                    size=6,
                    line=dict(
                        # color="#66CC99",
                        color=["red" if v > outStockRate else "green" for v in result_df['断货率']],
                        width=2
                    )
                ),
                # 正确提取单行对应的断货率文本
                text=[f"{v * 100:.1f}%" if v == v else "" for v in result_df['断货率']],
                textposition="top center",
                textfont=dict(
                    # color="#009966",
                    color=["red" if v > outStockRate else "green" for v in result_df['断货率']],

                    size=12,
                    family="Microsoft YaHei"
                ),
                showlegend=True,
                hoverinfo='skip',
                name='断货率'
            ),
            row=2, col=1, secondary_y=False
        )
    market_counts = result_df['子市场'].value_counts(sort=False)
    unique_markets = result_df['子市场'].unique()
    shapes = []
    current_position = -0.5
    for market in unique_markets[:-1]:
        current_position += market_counts[market]
        shapes.append(
            dict(
                type="line",
                xref="x2",
                yref="paper",
                x0=current_position,
                x1=current_position,
                y0=0,
                y1=1,
                line=dict(
                    color="#CCCCCC",
                    width=2,
                    dash="dash"
                )
            )
        )

    # ==================== 样式与坐标轴更新 ====================
    fig_子市场指标.update_layout(
        barmode='stack',
        height=650,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.05,
            xanchor="center",
            x=0.5,
            font=dict(size=14, color="black")
        ),
        plot_bgcolor='white',
        margin=dict(t=40, b=10, l=10, r=10),
        font=dict(family="Microsoft YaHei"),
        shapes=shapes
    )

    fig_子市场指标.update_xaxes(visible=False, row=1, col=1)
    fig_子市场指标.update_yaxes(visible=False, range=[-0.05, 1.15], row=1, col=1, secondary_y=False)
    fig_子市场指标.update_yaxes(showgrid=False, zeroline=False, range=[0, (
                result_df['海外在库周转'].fillna(0) + result_df['海外在途周转'].fillna(0)).max() * 1.3], row=2,
                                col=1, secondary_y=False)

    # 下方柱状图 X 轴
    fig_子市场指标.update_xaxes(
        tickangle=60,
        showdividers=True,
        dividercolor="#999999",
        row=2, col=1,
        tickfont=dict(size=12, color="black")
    )
    return fig_子市场指标




@st.fragment
def delivery_stock_area(df_海外周转, df_断货率):
    df_hw = df_海外周转.groupby(["周数",'子市场','层级']).agg(
         期初库存金额 = ("期初库存金额", "sum"),
         期末库存金额 = ("期末库存金额", "sum"),
         期初在途金额 = ("期初在途金额", "sum"),
         期末在途金额 = ("期末在途金额", "sum"),
         周销售出库金额 = ("周销售出库金额", "sum"),
         SKU数量 = ("SKU数量", "sum"),
    ).reset_index()
    df_hw['海外在库周转'] = ((df_hw['期末库存金额'] + df_hw['期初库存金额'])/2 / (df_hw['周销售出库金额']/7+0.001)).round(2)
    df_hw['海外在途周转'] = ((df_hw['期末在途金额'] + df_hw['期初在途金额'])/2 / (df_hw['周销售出库金额']/7+0.001)).round(2)
    # 合并结果
    df_nostock = df_断货率.groupby(["周数",'子市场','层级']).agg(
        SKU总数 = ("sku总数", "sum"),
        断货SKU数量 = ("断货sku数量", "sum"),
    ).reset_index()
    df_nostock['断货率'] = df_nostock['断货SKU数量']/df_nostock['SKU总数'].round(4)
    # df_nostock['断货率'] = df_nostock['断货率']*100
    result_df = pd.merge(df_hw, df_nostock[['周数','子市场','层级','断货率']], on=['周数','子市场','层级'], how='left')
    result_df['海外在库周转'] = np.where(result_df['海外在库周转']>1000, 0.001, result_df['海外在库周转'])
    result_df['海外在途周转'] = np.where(result_df['海外在途周转']>1000, 0.001, result_df['海外在途周转'])
    result_df['海外周转天数'] = (((result_df['海外在库周转'] + result_df['海外在途周转']).round(1)))

    result_df=result_df[result_df['海外周转天数']>0]
    result_df=result_df[~result_df['子市场'].isin(['LC-MX','LCM-MX','CP-JCW','MC-MX'])]
    result_df = result_df.reset_index(drop=True)
    six_weeks = result_df['周数'].sort_values().unique()[-4:]
    result_df = result_df[result_df['周数'].isin(six_weeks)]
    fig_子市场指标_TOP0 = 子市场指标分层级(result_df[result_df['层级']=='TOP0'],90,90,0.01)
    fig_子市场指标_TOP1 = 子市场指标分层级(result_df[result_df['层级']=='TOP1'],90,105,0.02)
    fig_子市场指标_TOP10 = 子市场指标分层级(result_df[result_df['层级']=='TOP10'],90,105,0.04)
    fig_子市场指标_TOP20 = 子市场指标分层级(result_df[result_df['层级']=='TOP20'],90,105,0.06)
    fig_子市场指标_普通 = 子市场指标分层级(result_df[result_df['层级']=='普通'],90,105,0.074)
    return fig_子市场指标_TOP0, fig_子市场指标_TOP1, fig_子市场指标_TOP10, fig_子市场指标_TOP20, fig_子市场指标_普通

def 整体堆积柱状图(df_海外周转, df_国内在库周转, df_断货率, targetDays1, targetDays2, outStockRate):
    if df_海外周转 is None or df_海外周转.empty:
        st.warning("无数据")    
        return

    df_hw = df_海外周转.groupby(['周数','层级']).agg(
         期初库存金额 = ("期初库存金额", "sum"),
         期末库存金额 = ("期末库存金额", "sum"),
         期初在途金额 = ("期初在途金额", "sum"),
         期末在途金额 = ("期末在途金额", "sum"),
         周销售出库金额 = ("周销售出库金额", "sum"),
    ).reset_index()
    df_hw['海外在库周转'] = ((df_hw['期末库存金额'] + df_hw['期初库存金额'])/2 / (df_hw['周销售出库金额']/7+0.001)).round(2)
    df_hw['海外在途周转'] = ((df_hw['期末在途金额'] + df_hw['期初在途金额'])/2 / (df_hw['周销售出库金额']/7+0.001)).round(2)
    # 合并结果
    df_nostock = df_断货率.groupby(['周数']).agg(
        SKU总数 = ("sku总数", "sum"),
        断货SKU数量 = ("断货sku数量", "sum"),
    ).reset_index()
    df_nostock['断货率'] = df_nostock['断货SKU数量']/df_nostock['SKU总数'].round(4)
    df_nostock['断货率'] = df_nostock['断货率']*100
    result_df = pd.merge(df_国内在库周转[['周数','层级','国内在库周转天数']], df_hw, on=['周数','层级'], how='left')
    df_hw_all = df_海外周转.groupby(['周数']).agg(
         期初库存金额 = ("期初库存金额", "sum"),
         期末库存金额 = ("期末库存金额", "sum"),
         期初在途金额 = ("期初在途金额", "sum"),
         期末在途金额 = ("期末在途金额", "sum"),
         周销售出库金额 = ("周销售出库金额", "sum"),
    ).reset_index()
    df_hw_all['海外在库周转'] = ((df_hw_all['期末库存金额'] + df_hw_all['期初库存金额'])/2 / (df_hw_all['周销售出库金额']/7+0.001)).round(2)
    df_hw_all['海外在途周转'] = ((df_hw_all['期末在途金额'] + df_hw_all['期初在途金额'])/2 / (df_hw_all['周销售出库金额']/7+0.001)).round(2)
    df_hw_all['海外周转天数'] = (((df_hw_all['海外在库周转'] + df_hw_all['海外在途周转']).round(1)))
    df_国内在库周转_all = df_国内在库周转.groupby(['周数']).agg(
        周初库存金额=("周初库存金额", "sum"),
        周末库存金额=("周末库存金额", "sum"),
        周销售出库金额=("周销售出库金额", "sum")
    ).reset_index()
    df_国内在库周转_all['国内在库周转天数'] = (df_国内在库周转_all['周初库存金额'] + df_国内在库周转_all['周末库存金额'])/2 / (df_国内在库周转_all['周销售出库金额']/7+0.001).round(2)

    result_df_all = pd.merge(df_国内在库周转_all[['周数','国内在库周转天数']], df_hw_all, on=['周数'], how='left')
    result_df_all = pd.merge(result_df_all, df_nostock, on=['周数'], how='left')
    result_df_all['总周转天数'] = result_df_all['海外周转天数'] + result_df_all['国内在库周转天数']
    result_df = result_df.sort_values(by=['周数','层级'])
    result_df = result_df.reset_index(drop=True)
    result_df_all = result_df_all.sort_values(by=['周数'])
    result_df_all = result_df_all.reset_index(drop=True)
    fig_历史周转 = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.1,
        row_heights=[0.3, 0.7]
    )
    color_map = {
        '海外在库': ['#2E86C1', '#5DADE2', '#85C1E9', '#AED6F1', '#D6EAF8'], # 蓝色系
        '海外在途': ['#E67E22', '#F39C12', '#F8C471', '#FAD7A0', '#FDEBD0'], # 橙色系
        '国内在库': ['#27AE60', '#52BE80', '#82E0AA', '#ABEBC6', '#D5F5E3']  # 绿色系
    }

    unique_levels = result_df['层级'].unique()
    for i, level in enumerate(unique_levels):
        level_data = result_df[result_df['层级'] == level]
        fig_历史周转.add_trace(go.Bar(
            x=level_data['周数'],
            y=level_data['海外在库周转'],
            name=f'{level}-海外在库',
            offsetgroup='1',
            legendgroup='海外在库',
            marker_color=color_map['海外在库'][i % 5],
            text=level_data['海外在库周转'].apply(lambda x: f"{x:.0f}" if x > 5 else ""),
            textposition='inside',
            hovertemplate="周: %{x}<br>层级: " + level + "<br>海外在库: %{y}天<extra></extra>"
        ), row=2, col=1)

        # 系列3：海外在途 (offsetgroup='3')
        fig_历史周转.add_trace(go.Bar(
            x=level_data['周数'],
            y=level_data['海外在途周转'],
            name=f'{level}-海外在途',
            offsetgroup='2',
            legendgroup='海外在途',
            marker_color=color_map['海外在途'][i % 5],
            text=level_data['海外在途周转'].apply(lambda x: f"{x:.0f}" if x > 5 else ""),
            textposition='inside',
            hovertemplate="周: %{x}<br>层级: " + level + "<br>海外在途: %{y}天<extra></extra>"
        ), row=2, col=1)
        fig_历史周转.add_trace(go.Bar(
            x=level_data['周数'],
            y=level_data['国内在库周转天数'],
            name=f'{level}-国内在库',
            offsetgroup='3',
            legendgroup='国内',
            marker_color=color_map['国内在库'][i % 5],
            text=level_data['国内在库周转天数'].apply(lambda x: f"{x:.0f}" if x > 5 else ""),
            textposition='inside',
            hovertemplate="周: %{x}<br>层级: " + level + "<br>国内在库: %{y}天<extra></extra>"
        ), row=2, col=1)

    stack_heights = result_df.groupby('周数').agg({
        '国内在库周转天数': 'sum',
        '海外在库周转': 'sum',
        '海外在途周转': 'sum'
    }).max(axis=1).reset_index()

    stack_heights.columns = ['周数', '柱子最高点']

    # 2. 将最高点合并到 result_df_all 中
    result_df_all = pd.merge(result_df_all, stack_heights, on='周数', how='left')
    fig_历史周转.add_trace(go.Scatter(
        x=result_df_all['周数'],
        y=result_df_all['柱子最高点'] + 30, 
        mode='lines+markers+text',
        name='总周转天数',
        text=result_df_all['总周转天数'].apply(lambda x: f"{x:.1f}"),
        textposition="top center",
        connectgaps=True,
        line=dict(
            color='#2E8421',
            width=3,
        ),
        marker=dict(
            color='#2E8421',
            size=12,
            line=dict(width=2, color='white')
        ),
        textfont=dict(
            size=15,
            color="black"
        ),
        # 悬浮窗显示真实数据
        hovertemplate="周数: %{x}<br>总周转天数: %{text}<extra></extra>"
    ),
        row=2,
        col=1
    )
    
    
    def get_trend_text_and_color(val):
        if val > outStockRate:
            return f"{val:.2f}%", "red"
        else:
            return f"{abs(val):.2f}%", "green"

    text_labels = result_df_all['断货率'].apply(lambda x: get_trend_text_and_color(x)[0])
    marker_colors = result_df_all['断货率'].apply(lambda x: get_trend_text_and_color(x)[1])

    fig_历史周转.add_trace(
        go.Scatter(
            x=result_df_all['周数'],
            y=result_df_all['断货率'],
            mode='markers+text',
            text=text_labels,
            textposition="top center",
            textfont=dict(color=marker_colors, size=16),
            marker=dict(
                color=marker_colors,
                size=15,
                line=dict(width=1, color='white')
            ),
            error_y=dict(
                type='data',
                symmetric=False,
                array=[0] * len(result_df_all),  
                arrayminus=result_df_all['断货率'],
                width=0,
                thickness=1.5,
                color='rgba(100, 100, 100, 0.5)'
            ),
            showlegend=True,
            name='断货率',
            hovertemplate='周数: %{x}<br>断货率: %{y}%<extra></extra>'
        ),
        row=1, col=1
    )
    max_stack_height = max(result_df_all['柱子最高点']) + 150
    fig_历史周转.update_layout(
        barmode='stack',
        height=800,
        plot_bgcolor='white',
        margin=dict(t=80, b=50, l=50, r=50),
        font=dict(family="Microsoft YaHei"),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5,
            title_text="",
            font=dict(size=16, color="black")
        )
    )

    # max_y = max(result_df_all['总周转天数']) + 300
    fig_历史周转.update_yaxes(
        title_text="周转天数",
        showgrid=True,
        gridcolor='lightgray',
        row=2, col=1,
        tickfont=dict(size=16, color="black"),
        range=[0, max_stack_height]
    )
    max_out_of_stock = result_df_all['断货率'].max()
    upper_limit = max_out_of_stock * 1.5 if max_out_of_stock > 0 else 1 
    fig_历史周转.update_yaxes(
        range=[0, upper_limit],
        showticklabels=False,
        showgrid=False,
        zeroline=True,
        zerolinecolor='gray',
        zerolinewidth=1,
        row=1, col=1
    )

    fig_历史周转.update_xaxes(
        type='category',
        row=2, col=1,
        tickfont=dict(size=16, color="black")
    )
    st.plotly_chart(fig_历史周转, width='stretch',key="fig_历史周转")



if st.session_state.df_海外周转 is not None:
    # 历史实际周转区域
    st.header("📈 历史实际周转", anchor="0")
    整体堆积柱状图(st.session_state.df_海外周转, st.session_state.df_国内在库周转, st.session_state.df_断货率,90,90,1)
    fig_历史周转_TOP0, fig_历史周转_TOP1, fig_历史周转_TOP10, fig_历史周转_TOP20, fig_历史周转_普通 = actual_turnover_area(st.session_state.df_海外周转, st.session_state.df_国内在库周转, st.session_state.df_断货率)
    fig_子市场指标_TOP0, fig_子市场指标_TOP1, fig_子市场指标_TOP10, fig_子市场指标_TOP20, fig_子市场指标_普通 = delivery_stock_area(st.session_state.df_海外周转, st.session_state.df_断货率)
    st.header("📈 TOP0", anchor="1")
    st.plotly_chart(fig_历史周转_TOP0, width='stretch',key="fig_历史周转_TOP0")
    st.plotly_chart(fig_子市场指标_TOP0, width='stretch',key="fig_子市场指标_TOP0")
    st.divider()
    st.header("📈 TOP1", anchor="1")
    st.plotly_chart(fig_历史周转_TOP1, width='stretch',key="fig_历史周转_TOP1")
    st.plotly_chart(fig_子市场指标_TOP1, width='stretch',key="fig_子市场指标_TOP1")
    st.divider()
    st.header("📈 TOP10", anchor="1")
    st.plotly_chart(fig_历史周转_TOP10, width='stretch',key="fig_历史周转_TOP10")
    st.plotly_chart(fig_子市场指标_TOP10, width='stretch',key="fig_子市场指标_TOP10")
    st.divider()
    st.header("📈 TOP20", anchor="1")
    st.plotly_chart(fig_历史周转_TOP20, width='stretch',key="fig_历史周转_TOP20")
    st.plotly_chart(fig_子市场指标_TOP20, width='stretch',key="fig_子市场指标_TOP20")
    st.divider()
    st.header("📈 普通", anchor="1")
    st.plotly_chart(fig_历史周转_普通, width='stretch',key="fig_历史周转_普通")
    st.plotly_chart(fig_子市场指标_普通, width='stretch',key="fig_子市场指标_普通")
    st.divider()
    
else:
    st.info("👋 请先在左侧侧边栏上传数据文件。")
